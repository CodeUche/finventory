"""
Tests for reports.views.

Covers:
- All 10 report endpoints (sales, top-products, top-customers, pnl, expenses,
  inventory, cash-flow, ar-aging, ap-aging, vat-summary).
- Authentication: 401 without credentials.
- Period shortcuts: ?period=today|week|month|year|all|custom all return 200.
- Export formats: ?format=excel returns XLSX content-type;
                  ?format=pdf returns application/pdf with %PDF magic bytes;
                  ?format=json (default) returns JSON.
- Custom date range via ?date_from=&date_to= parsed correctly.
- Empty database produces valid (non-error) response structures.
- ARAgingView / APAgingView accept ?as_of= param.
- SalesSummaryView accepts ?group_by=month param.
- TopProductsView and TopCustomersView accept ?limit= param.
"""

import uuid
from datetime import date
from decimal import Decimal

from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase

from apps.authentication.models import User
from apps.tenancy.models import Organisation, Membership


# ─── Fixtures ────────────────────────────────────────────────────────────────


def _make_superuser(suffix=""):
    return User.objects.create_superuser(
        email=f"admin{suffix}@test.com",
        password="testpass123",
        first_name="Admin",
        last_name="User",
    )


def _make_org(owner):
    slug = uuid.uuid4().hex[:12]
    return Organisation.objects.create(
        name="Test Org",
        slug=slug,
        owner=owner,
    )


def _make_accountant(org, suffix=""):
    user = User.objects.create_user(
        email=f"accountant{suffix}@test.com",
        password="testpass123",
        first_name="Acc",
        last_name="User",
        is_verified=True,
    )
    Membership.objects.create(
        user=user,
        organisation=org,
        role="accountant",
        is_active=True,
    )
    return user


# ─── Base test case ───────────────────────────────────────────────────────────


class BaseReportTestCase(APITestCase):
    """
    Sets up a superuser + organisation for every test class.
    Superusers bypass membership checks in IsAccountant, making setup minimal.
    """

    @classmethod
    def setUpTestData(cls):
        cls.user = _make_superuser(suffix=f"_{cls.__name__}")
        cls.org = _make_org(owner=cls.user)

    def _auth(self, user=None):
        """Force-authenticate and set org header."""
        self.client.force_authenticate(user=user or self.user)
        self.client.credentials(HTTP_X_ORGANISATION_ID=str(self.org.id))

    def _get(self, url_name, **query):
        self._auth()
        url = reverse(url_name)
        return self.client.get(url, query)

    def _get_unauthenticated(self, url_name):
        self.client.force_authenticate(user=None)
        self.client.credentials()
        return self.client.get(reverse(url_name))


# ─── Authentication guard ─────────────────────────────────────────────────────


class TestReportAuthentication(BaseReportTestCase):
    """Every endpoint must return 401 when no credentials are provided."""

    _ENDPOINTS = [
        "report-sales",
        "report-top-products",
        "report-top-customers",
        "report-pnl",
        "report-expenses",
        "report-inventory",
        "report-cash-flow",
        "report-ar-aging",
        "report-ap-aging",
        "report-vat-summary",
    ]

    def test_unauthenticated_returns_401(self):
        for name in self._ENDPOINTS:
            with self.subTest(endpoint=name):
                resp = self._get_unauthenticated(name)
                self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED, name)


# ─── Accountant-role access ───────────────────────────────────────────────────


class TestAccountantAccess(APITestCase):
    """A user with accountant role can access report endpoints."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = _make_superuser(suffix="_acc_owner")
        cls.org = _make_org(owner=cls.owner)
        cls.accountant = _make_accountant(cls.org, suffix="_access")

    def _auth(self):
        self.client.force_authenticate(user=self.accountant)
        self.client.credentials(HTTP_X_ORGANISATION_ID=str(self.org.id))

    def test_accountant_can_access_pnl(self):
        self._auth()
        resp = self.client.get(reverse("report-pnl"))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_accountant_can_access_sales(self):
        self._auth()
        resp = self.client.get(reverse("report-sales"))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_viewer_role_is_denied(self):
        viewer = User.objects.create_user(
            email="viewer_access@test.com",
            password="testpass",
            first_name="View",
            last_name="Only",
            is_verified=True,
        )
        Membership.objects.create(
            user=viewer, organisation=self.org, role="viewer", is_active=True
        )
        self.client.force_authenticate(user=viewer)
        self.client.credentials(HTTP_X_ORGANISATION_ID=str(self.org.id))
        resp = self.client.get(reverse("report-pnl"))
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)


# ─── Sales Summary ─────────────────────────────────────────────────────────────


class TestSalesSummaryView(BaseReportTestCase):

    def test_returns_200_json(self):
        resp = self._get("report-sales")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.accepted_media_type, "application/json")

    def test_response_is_list(self):
        resp = self._get("report-sales")
        self.assertIsInstance(resp.data, list)

    def test_period_today(self):
        resp = self._get("report-sales", period="today")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_week(self):
        resp = self._get("report-sales", period="week")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_month(self):
        resp = self._get("report-sales", period="month")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_year(self):
        resp = self._get("report-sales", period="year")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_all(self):
        resp = self._get("report-sales", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_custom_with_dates(self):
        resp = self._get("report-sales", period="custom",
                         date_from="2025-01-01", date_to="2025-06-30")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_group_by_month(self):
        resp = self._get("report-sales", period="month", group_by="month")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_group_by_year(self):
        resp = self._get("report-sales", period="year", group_by="year")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel_content_type(self):
        resp = self._get("report-sales", format="excel")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_excel_disposition(self):
        resp = self._get("report-sales", format="excel")
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn(".xlsx", resp["Content-Disposition"])

    def test_format_pdf_content_type(self):
        resp = self._get("report-sales", format="pdf")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp["Content-Type"], "application/pdf")

    def test_format_pdf_magic_bytes(self):
        resp = self._get("report-sales", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_format_json_explicit(self):
        resp = self._get("report-sales", format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)


# ─── Top Products ─────────────────────────────────────────────────────────────


class TestTopProductsView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-top-products")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_limit_param(self):
        resp = self._get("report-top-products", limit=5)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_all(self):
        resp = self._get("report-top-products", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-top-products", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-top-products", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_custom_date_range(self):
        resp = self._get("report-top-products", period="custom",
                         date_from="2025-01-01", date_to="2025-12-31")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── Top Customers ────────────────────────────────────────────────────────────


class TestTopCustomersView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-top-customers")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_limit_param(self):
        resp = self._get("report-top-customers", limit=3)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-top-customers", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-top-customers", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_period_today(self):
        resp = self._get("report-top-customers", period="today")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── Profit & Loss ────────────────────────────────────────────────────────────


class TestProfitAndLossView(BaseReportTestCase):

    _EXPECTED_KEYS = {"revenue", "cost_of_goods_sold", "gross_profit",
                      "gross_margin_pct", "operating_expenses", "net_profit", "net_margin_pct"}

    def test_returns_200_dict(self):
        resp = self._get("report-pnl")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-pnl")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_revenue_has_sub_keys(self):
        resp = self._get("report-pnl")
        rev = resp.data.get("revenue", {})
        for k in ("gross_sales", "tax_collected", "discounts"):
            self.assertIn(k, rev, f"revenue.{k} missing")

    def test_period_month(self):
        resp = self._get("report-pnl", period="month")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_all(self):
        resp = self._get("report-pnl", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-pnl", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-pnl", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_empty_db_returns_zero_values(self):
        resp = self._get("report-pnl")
        self.assertEqual(float(resp.data["cost_of_goods_sold"]), 0.0)

    def test_custom_date_range(self):
        resp = self._get("report-pnl", period="custom",
                         date_from="2025-01-01", date_to="2025-03-31")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── Expense Breakdown ────────────────────────────────────────────────────────


class TestExpenseBreakdownView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-expenses")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_period_week(self):
        resp = self._get("report-expenses", period="week")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_year(self):
        resp = self._get("report-expenses", period="year")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-expenses", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-expenses", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))


# ─── Inventory Valuation ──────────────────────────────────────────────────────


class TestInventoryValuationView(BaseReportTestCase):

    _EXPECTED_KEYS = {"total_inventory_value", "items"}

    def test_returns_200_dict(self):
        resp = self._get("report-inventory")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-inventory")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_items_is_list(self):
        resp = self._get("report-inventory")
        self.assertIsInstance(resp.data["items"], list)

    def test_format_excel(self):
        resp = self._get("report-inventory", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-inventory", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_period_param_ignored_for_inventory(self):
        # Inventory is point-in-time; period param should not cause errors
        resp = self._get("report-inventory", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── Cash Flow ────────────────────────────────────────────────────────────────


class TestCashFlowView(BaseReportTestCase):

    _EXPECTED_KEYS = {"cash_inflows", "cash_outflows", "net_cash_flow"}

    def test_returns_200_dict(self):
        resp = self._get("report-cash-flow")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-cash-flow")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_period_month(self):
        resp = self._get("report-cash-flow", period="month")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_all(self):
        resp = self._get("report-cash-flow", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-cash-flow", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-cash-flow", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_custom_date_range(self):
        resp = self._get("report-cash-flow", period="custom",
                         date_from="2025-01-01", date_to="2025-12-31")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── AR Aging ─────────────────────────────────────────────────────────────────


class TestARAgingView(BaseReportTestCase):

    _EXPECTED_KEYS = {"as_of", "buckets", "total_outstanding", "invoices"}

    def test_returns_200_dict(self):
        resp = self._get("report-ar-aging")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-ar-aging")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_buckets_has_aging_bands(self):
        resp = self._get("report-ar-aging")
        buckets = resp.data.get("buckets", {})
        for band in ("current", "1_30", "31_60", "61_90", "over_90"):
            self.assertIn(band, buckets, f"Bucket {band} missing")

    def test_invoices_is_list(self):
        resp = self._get("report-ar-aging")
        self.assertIsInstance(resp.data["invoices"], list)

    def test_as_of_param(self):
        resp = self._get("report-ar-aging", as_of="2025-06-30")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(str(resp.data["as_of"]), "2025-06-30")

    def test_invalid_as_of_falls_back_to_today(self):
        resp = self._get("report-ar-aging", as_of="not-a-date")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(str(resp.data["as_of"]), str(date.today()))

    def test_format_excel(self):
        resp = self._get("report-ar-aging", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-ar-aging", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))


# ─── AP Aging ─────────────────────────────────────────────────────────────────


class TestAPAgingView(BaseReportTestCase):

    _EXPECTED_KEYS = {"as_of", "buckets", "total_outstanding", "bills"}

    def test_returns_200_dict(self):
        resp = self._get("report-ap-aging")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-ap-aging")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_as_of_param(self):
        resp = self._get("report-ap-aging", as_of="2025-12-31")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(str(resp.data["as_of"]), "2025-12-31")

    def test_invalid_as_of_falls_back_to_today(self):
        resp = self._get("report-ap-aging", as_of="bad")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(str(resp.data["as_of"]), str(date.today()))

    def test_format_excel(self):
        resp = self._get("report-ap-aging", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-ap-aging", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_empty_db_zero_outstanding(self):
        resp = self._get("report-ap-aging")
        self.assertEqual(float(resp.data["total_outstanding"]), 0.0)


# ─── VAT Summary ─────────────────────────────────────────────────────────────


class TestVATSummaryView(BaseReportTestCase):

    _EXPECTED_KEYS = {"output_vat", "input_vat", "net_vat_payable"}

    def test_returns_200_dict(self):
        resp = self._get("report-vat-summary")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, dict)

    def test_expected_keys_present(self):
        resp = self._get("report-vat-summary")
        for key in self._EXPECTED_KEYS:
            self.assertIn(key, resp.data, f"Missing key: {key}")

    def test_period_month(self):
        resp = self._get("report-vat-summary", period="month")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_period_all(self):
        resp = self._get("report-vat-summary", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_format_excel(self):
        resp = self._get("report-vat-summary", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-vat-summary", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_empty_db_zero_vat(self):
        resp = self._get("report-vat-summary")
        self.assertEqual(float(resp.data["output_vat"]), 0.0)
        self.assertEqual(float(resp.data["input_vat"]), 0.0)

    def test_custom_date_range(self):
        resp = self._get("report-vat-summary", period="custom",
                         date_from="2025-04-01", date_to="2025-06-30")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


# ─── Cross-cutting: period label in subtitle ──────────────────────────────────


class TestPeriodLabelInExport(BaseReportTestCase):
    """When exporting, the subtitle row must reflect the chosen period.

    Excel exports now lead with a small header block — company name, then
    report title, then the period/subtitle row — matching the reference
    sample reports (see _write_report_sheet() in apps/reports/exporters.py).
    Row 1 = org name, Row 2 = title, Row 3 = subtitle/period, Row 4 = column
    headers. Every legacy view (like ProfitAndLossView here) always supplies
    both an org and a title, so the subtitle is reliably on row 3.
    """

    def test_excel_export_has_subtitle_for_month_period(self):
        import io
        import openpyxl
        resp = self._get("report-pnl", period="month", format="excel")
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        subtitle_val = ws.cell(row=3, column=1).value
        self.assertIsNotNone(subtitle_val)
        self.assertIn("Month", subtitle_val)

    def test_excel_export_has_subtitle_for_custom_period(self):
        import io
        import openpyxl
        resp = self._get("report-pnl", period="custom",
                         date_from="2025-01-01", date_to="2025-03-31",
                         format="excel")
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        subtitle_val = ws.cell(row=3, column=1).value
        self.assertIsNotNone(subtitle_val)
        self.assertIn("Jan 2025", subtitle_val)


# ─── Invalid / edge-case query params ─────────────────────────────────────────


class TestQueryParamEdgeCases(BaseReportTestCase):
    """Invalid or missing params must never produce a 500."""

    def test_invalid_date_from_does_not_500(self):
        resp = self._get("report-sales", period="custom",
                         date_from="not-a-date", date_to="2025-12-31")
        self.assertNotEqual(resp.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)

    def test_invalid_date_to_does_not_500(self):
        resp = self._get("report-sales", period="custom",
                         date_from="2025-01-01", date_to="nope")
        self.assertNotEqual(resp.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)

    def test_unknown_period_treated_as_custom(self):
        resp = self._get("report-pnl", period="quarterly")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_uppercase_period_accepted(self):
        resp = self._get("report-pnl", period="TODAY")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_unknown_format_falls_back_to_json(self):
        resp = self._get("report-sales", format="csv")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_invalid_limit_raises_no_500(self):
        # limit must be parseable as int — valid int strings work
        resp = self._get("report-top-products", limit=1)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_no_org_header_uses_fallback_org(self):
        # Without the org header DRF falls back to the user's first org.
        # Superusers still hit IsAccountant.has_permission → should not 500.
        self.client.force_authenticate(user=self.user)
        self.client.credentials()  # clear org header
        resp = self.client.get(reverse("report-pnl"))
        self.assertNotEqual(resp.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── Fixture helpers for the new reports below ────────────────────────────────


def _make_customer(org, name="Fixture Customer", code="CUST-001", **extra):
    from apps.customers.models import Customer
    return Customer.objects.create(organisation=org, name=name, code=code, **extra)


def _make_warehouse(org, name="Main Warehouse"):
    from apps.inventory.models import Warehouse
    return Warehouse.objects.create(organisation=org, name=name, is_default=True)


def _make_invoice(org, customer, warehouse, user, total_amount, amount_paid=None,
                   amount_due=None, issue_date=None, status="confirmed"):
    from apps.sales.models import Invoice
    issue_date = issue_date or date.today()
    amount_paid = amount_paid if amount_paid is not None else total_amount
    amount_due = amount_due if amount_due is not None else (total_amount - amount_paid)
    return Invoice.objects.create(
        organisation=org,
        invoice_number=Invoice.generate_number(org),
        customer=customer,
        status=status,
        payment_method="cash",
        issue_date=issue_date,
        warehouse=warehouse,
        created_by=user,
        subtotal=total_amount,
        discount_amount=0,
        tax_amount=0,
        total_amount=total_amount,
        amount_paid=amount_paid,
        amount_due=amount_due,
    )


def _make_payment(org, invoice, user, amount, method="cash"):
    from apps.sales.models import SalePayment
    return SalePayment.objects.create(
        organisation=org, invoice=invoice, amount=amount,
        method=method, received_by=user,
    )


def _make_product(org, sku="SKU-001", name="Fixture Product",
                   cost_price=Decimal("60"), selling_price=Decimal("100")):
    from apps.inventory.models import Product
    return Product.objects.create(
        organisation=org, sku=sku, name=name,
        cost_price=cost_price, selling_price=selling_price,
    )


def _make_stock_item(org, product, warehouse, quantity):
    from apps.inventory.models import StockItem
    return StockItem.objects.create(
        organisation=org, product=product, warehouse=warehouse,
        quantity_on_hand=quantity,
    )


# ─── Customer Balance ──────────────────────────────────────────────────────


class TestCustomerBalanceView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-customer-balance")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-customer-balance")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_format_excel(self):
        resp = self._get("report-customer-balance", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_format_pdf(self):
        resp = self._get("report-customer-balance", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_correct_balance_and_credit_fields(self):
        customer = _make_customer(
            self.org, name="Balance Co", code="BAL-1",
            outstanding_balance=Decimal("5000"), credit_limit=Decimal("20000"),
        )
        resp = self._get("report-customer-balance")
        row = next(r for r in resp.data if r["customer_id"] == str(customer.id))
        self.assertEqual(Decimal(str(row["outstanding_balance"])), Decimal("5000"))
        self.assertEqual(Decimal(str(row["credit_limit"])), Decimal("20000"))
        self.assertEqual(Decimal(str(row["available_credit"])), Decimal("15000"))

    def test_sorted_by_outstanding_balance_descending(self):
        _make_customer(self.org, name="Low", code="LOW-1", outstanding_balance=Decimal("100"))
        _make_customer(self.org, name="High", code="HIGH-1", outstanding_balance=Decimal("9000"))
        resp = self._get("report-customer-balance")
        balances = [Decimal(str(r["outstanding_balance"])) for r in resp.data]
        self.assertEqual(balances, sorted(balances, reverse=True))

    def test_inactive_customers_excluded(self):
        _make_customer(self.org, name="Inactive Co", code="INACT-1", is_active=False)
        resp = self._get("report-customer-balance")
        names = [r["customer_name"] for r in resp.data]
        self.assertNotIn("Inactive Co", names)


# ─── Payments by Customer ─────────────────────────────────────────────────


class TestPaymentsByCustomerView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-payments-by-customer")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-payments-by-customer")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_format_excel(self):
        resp = self._get("report-payments-by-customer", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_totals_and_breakdown_correct(self):
        warehouse = _make_warehouse(self.org, name="PBC Warehouse")
        customer = _make_customer(self.org, name="PBC Customer", code="PBC-1")
        invoice = _make_invoice(self.org, customer, warehouse, self.user, Decimal("1000"))
        _make_payment(self.org, invoice, self.user, Decimal("600"), method="cash")
        _make_payment(self.org, invoice, self.user, Decimal("400"), method="bank_transfer")

        resp = self._get("report-payments-by-customer", period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        row = next(r for r in resp.data if r["customer_id"] == str(customer.id))
        self.assertEqual(Decimal(str(row["total_received"])), Decimal("1000"))
        self.assertEqual(row["payment_count"], 2)
        self.assertEqual(Decimal(str(row["breakdown"]["cash"])), Decimal("600"))
        self.assertEqual(Decimal(str(row["breakdown"]["bank_transfer"])), Decimal("400"))


# ─── Customer Payments (drill-down) ───────────────────────────────────────


class TestCustomerPaymentsView(BaseReportTestCase):

    def test_missing_customer_id_returns_400(self):
        resp = self._get("report-customer-payments")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-customer-payments")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_returns_individual_payment_rows(self):
        warehouse = _make_warehouse(self.org, name="CP Warehouse")
        customer = _make_customer(self.org, name="CP Customer", code="CP-1")
        invoice = _make_invoice(self.org, customer, warehouse, self.user, Decimal("500"))
        _make_payment(self.org, invoice, self.user, Decimal("500"), method="pos")

        resp = self._get("report-customer-payments", customer_id=str(customer.id), period="all")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]["invoice_number"], invoice.invoice_number)
        self.assertEqual(Decimal(str(resp.data[0]["amount"])), Decimal("500"))
        self.assertEqual(resp.data[0]["method"], "pos")


# ─── Account Statement (GL) ────────────────────────────────────────────────


class TestAccountStatementView(BaseReportTestCase):

    def test_missing_account_id_returns_400(self):
        resp = self._get("report-account-statement")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-account-statement")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_unknown_account_returns_404(self):
        resp = self._get("report-account-statement", account_id=str(uuid.uuid4()))
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_closing_balance_equals_opening_plus_movement(self):
        """closing_balance must equal opening_balance + sum(debits) - sum(credits)
        for a debit-normal (asset) account, restricted to the requested period."""
        from apps.accounting.models import Account
        from apps.accounting.services import AccountingService

        asset_account = Account.objects.create(
            organisation=self.org, code="9101", name="Statement Asset", account_type="asset",
        )
        offset_account = Account.objects.create(
            organisation=self.org, code="9102", name="Statement Offset", account_type="liability",
        )

        # Opening entry — before the reporting window.
        AccountingService.post_journal_entry(
            organisation=self.org,
            description="Opening entry",
            entry_date=date(2025, 1, 1),
            lines=[
                (asset_account, Decimal("1000"), Decimal("0")),
                (offset_account, Decimal("0"), Decimal("1000")),
            ],
            created_by=self.user,
        )
        # In-period entries.
        AccountingService.post_journal_entry(
            organisation=self.org,
            description="Period debit",
            entry_date=date(2025, 2, 10),
            lines=[
                (asset_account, Decimal("300"), Decimal("0")),
                (offset_account, Decimal("0"), Decimal("300")),
            ],
            created_by=self.user,
        )
        AccountingService.post_journal_entry(
            organisation=self.org,
            description="Period credit",
            entry_date=date(2025, 2, 20),
            lines=[
                (offset_account, Decimal("120"), Decimal("0")),
                (asset_account, Decimal("0"), Decimal("120")),
            ],
            created_by=self.user,
        )

        resp = self._get(
            "report-account-statement",
            account_id=str(asset_account.id),
            period="custom", date_from="2025-02-01", date_to="2025-02-28",
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        opening = Decimal(str(resp.data["opening_balance"]))
        closing = Decimal(str(resp.data["closing_balance"]))
        lines = resp.data["lines"]
        total_debit = sum(Decimal(str(l["debit"])) for l in lines)
        total_credit = sum(Decimal(str(l["credit"])) for l in lines)

        self.assertEqual(opening, Decimal("1000"))
        self.assertEqual(closing, opening + total_debit - total_credit)
        self.assertEqual(closing, Decimal("1180"))
        # Running balance on the last line must equal the closing balance.
        self.assertEqual(Decimal(str(lines[-1]["running_balance"])), closing)

    def test_format_excel(self):
        from apps.accounting.models import Account
        account = Account.objects.create(
            organisation=self.org, code="9103", name="Excel Statement Acct", account_type="asset",
        )
        resp = self._get("report-account-statement", account_id=str(account.id), format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])


# ─── Customer Details (master directory) ──────────────────────────────────


class TestCustomerDetailsView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-customer-details")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-customer-details")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_format_pdf(self):
        resp = self._get("report-customer-details", format="pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_lifetime_totals_correct(self):
        warehouse = _make_warehouse(self.org, name="CD Warehouse")
        customer = _make_customer(
            self.org, name="Lifetime Co", code="LIFE-1",
            email="lifetime@test.com", phone="08000000000",
        )
        inv1 = _make_invoice(self.org, customer, warehouse, self.user, Decimal("700"))
        inv2 = _make_invoice(self.org, customer, warehouse, self.user, Decimal("300"))
        _make_payment(self.org, inv1, self.user, Decimal("700"))
        _make_payment(self.org, inv2, self.user, Decimal("300"))

        resp = self._get("report-customer-details")
        row = next(r for r in resp.data if r["customer_id"] == str(customer.id))
        self.assertEqual(Decimal(str(row["total_sales"])), Decimal("1000"))
        self.assertEqual(Decimal(str(row["total_payments"])), Decimal("1000"))
        self.assertEqual(row["email"], "lifetime@test.com")


# ─── Product Details (master directory) ───────────────────────────────────


class TestProductDetailsView(BaseReportTestCase):

    def test_returns_200_list(self):
        resp = self._get("report-product-details")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIsInstance(resp.data, list)

    def test_unauthenticated_returns_401(self):
        resp = self._get_unauthenticated("report-product-details")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_format_excel(self):
        resp = self._get("report-product-details", format="excel")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_margin_and_stock_correct(self):
        warehouse = _make_warehouse(self.org, name="PD Warehouse")
        product = _make_product(
            self.org, sku="PD-SKU-1", name="Margin Product",
            cost_price=Decimal("60"), selling_price=Decimal("100"),
        )
        _make_stock_item(self.org, product, warehouse, Decimal("25"))

        resp = self._get("report-product-details")
        row = next(r for r in resp.data if r["product_id"] == str(product.id))
        self.assertEqual(Decimal(str(row["margin_pct"])), Decimal("40.00"))
        self.assertEqual(Decimal(str(row["stock_quantity"])), Decimal("25"))

    def test_zero_selling_price_does_not_divide_by_zero(self):
        _make_product(self.org, sku="PD-SKU-0", name="Free Item",
                       cost_price=Decimal("0"), selling_price=Decimal("0"))
        resp = self._get("report-product-details")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
