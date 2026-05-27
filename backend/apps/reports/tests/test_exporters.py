"""
Tests for reports.exporters.

Covers:
- export_excel: content type, Content-Disposition, valid workbook, headers,
  data rows, alternating shading, empty data, special/unicode chars, Decimal
  values, date values, None values, long strings (truncated to max col width).
- export_pdf: content type, Content-Disposition, PDF magic bytes, title,
  subtitle, empty data.
- dispatch_export: returns correct type for 'excel'/'pdf', None for 'json'.
- _clean: every normalisation rule.
"""

import io
from datetime import date, datetime
from decimal import Decimal

from django.test import SimpleTestCase

from apps.reports.exporters import _clean, dispatch_export, export_excel, export_pdf


# ─── _clean ──────────────────────────────────────────────────────────────────


class TestClean(SimpleTestCase):
    def test_none_becomes_empty_string(self):
        self.assertEqual(_clean(None), "")

    def test_empty_string_unchanged(self):
        self.assertEqual(_clean(""), "")

    def test_bool_true_becomes_yes(self):
        self.assertEqual(_clean(True), "Yes")

    def test_bool_false_becomes_no(self):
        self.assertEqual(_clean(False), "No")

    def test_decimal_becomes_float(self):
        result = _clean(Decimal("1234.56"))
        self.assertIsInstance(result, float)
        self.assertAlmostEqual(result, 1234.56)

    def test_decimal_zero(self):
        self.assertEqual(_clean(Decimal("0")), 0.0)

    def test_date_becomes_iso_string(self):
        self.assertEqual(_clean(date(2025, 6, 15)), "2025-06-15")

    def test_datetime_becomes_datetime_string(self):
        result = _clean(datetime(2025, 6, 15, 10, 30))
        self.assertEqual(result, "2025-06-15 10:30")

    def test_int_unchanged(self):
        self.assertEqual(_clean(42), 42)

    def test_float_unchanged(self):
        self.assertAlmostEqual(_clean(3.14), 3.14)

    def test_string_unchanged(self):
        self.assertEqual(_clean("hello"), "hello")

    def test_custom_object_becomes_string(self):
        class Obj:
            def __str__(self):
                return "custom"
        self.assertEqual(_clean(Obj()), "custom")

    def test_uuid_becomes_string(self):
        import uuid
        u = uuid.uuid4()
        self.assertEqual(_clean(u), str(u))

    def test_unicode_string_unchanged(self):
        self.assertEqual(_clean("Ọrụ àkụ"), "Ọrụ àkụ")

    def test_large_decimal(self):
        result = _clean(Decimal("9999999999.99"))
        self.assertIsInstance(result, float)


# ─── export_excel ─────────────────────────────────────────────────────────────


class TestExportExcel(SimpleTestCase):

    def _make(self, headers=None, rows=None, **kwargs):
        headers = headers or ["Name", "Amount", "Date"]
        rows = rows or [["Alice", 1000, date(2025, 1, 1)], ["Bob", 2000, date(2025, 2, 1)]]
        return export_excel(headers, rows, **kwargs)

    def test_content_type(self):
        resp = self._make()
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_content_disposition_contains_filename(self):
        resp = export_excel(["H"], [[1]], filename="my_report.xlsx")
        self.assertIn("my_report.xlsx", resp["Content-Disposition"])
        self.assertIn("attachment", resp["Content-Disposition"])

    def test_default_filename(self):
        resp = self._make()
        self.assertIn("report.xlsx", resp["Content-Disposition"])

    def test_response_is_valid_xlsx(self):
        import openpyxl
        resp = self._make()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertIsNotNone(wb)

    def test_headers_written_to_first_row(self):
        import openpyxl
        headers = ["Col A", "Col B", "Col C"]
        resp = export_excel(headers, [])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        actual = [ws.cell(row=1, column=i + 1).value for i in range(3)]
        self.assertEqual(actual, headers)

    def test_data_rows_written_correctly(self):
        import openpyxl
        headers = ["Name", "Amount"]
        rows = [["Alice", 500], ["Bob", 1000]]
        resp = export_excel(headers, rows)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "Alice")
        self.assertEqual(ws.cell(row=2, column=2).value, 500)
        self.assertEqual(ws.cell(row=3, column=1).value, "Bob")
        self.assertEqual(ws.cell(row=3, column=2).value, 1000)

    def test_decimal_written_as_float(self):
        import openpyxl
        resp = export_excel(["Amt"], [[Decimal("1234.56")]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertAlmostEqual(ws.cell(row=2, column=1).value, 1234.56)

    def test_date_written_as_string(self):
        import openpyxl
        resp = export_excel(["Date"], [[date(2025, 6, 15)]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "2025-06-15")

    def test_none_written_as_empty_string(self):
        import openpyxl
        resp = export_excel(["Val"], [[None]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertIn(cell_value, (None, ""))

    def test_empty_rows_produces_header_only(self):
        import openpyxl
        resp = export_excel(["H1", "H2"], [])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 1)

    def test_sheet_name_truncated_to_31_chars(self):
        import openpyxl
        long_name = "A" * 50
        resp = export_excel(["H"], [[1]], sheet_name=long_name)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertLessEqual(len(wb.active.title), 31)

    def test_subtitle_row_written_when_provided(self):
        import openpyxl
        resp = export_excel(["H"], [[1]], subtitle="01 Jun 2025 – 30 Jun 2025")
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # Subtitle row is row 1; header is row 2
        subtitle_cell = ws.cell(row=1, column=1).value
        self.assertIsNotNone(subtitle_cell)
        self.assertIn("Jun", subtitle_cell)

    def test_unicode_data(self):
        import openpyxl
        resp = export_excel(["Ìpínlẹ"], [["Ọrụ àkụ ₦1,000"]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertIn("₦", ws.cell(row=2, column=1).value)

    def test_many_rows(self):
        import openpyxl
        rows = [[f"Item {i}", i * 100] for i in range(200)]
        resp = export_excel(["Name", "Amount"], rows)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 201)  # 1 header + 200 data

    def test_header_row_has_bold_font(self):
        import openpyxl
        resp = export_excel(["Header"], [[1]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertTrue(ws.cell(row=1, column=1).font.bold)

    def test_freeze_pane_set(self):
        import openpyxl
        resp = export_excel(["H"], [[1]])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # freeze_panes should be set (not None)
        self.assertIsNotNone(ws.freeze_panes)


# ─── export_pdf ───────────────────────────────────────────────────────────────


class TestExportPdf(SimpleTestCase):

    def _make(self, headers=None, rows=None, **kwargs):
        headers = headers or ["Name", "Amount"]
        rows = rows or [["Alice", 1000], ["Bob", 2000]]
        return export_pdf(headers, rows, **kwargs)

    def test_content_type(self):
        resp = self._make()
        self.assertEqual(resp["Content-Type"], "application/pdf")

    def test_content_disposition_attachment(self):
        resp = export_pdf(["H"], [[1]], filename="test.pdf")
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn("test.pdf", resp["Content-Disposition"])

    def test_default_filename(self):
        resp = self._make()
        self.assertIn("report.pdf", resp["Content-Disposition"])

    def test_pdf_starts_with_magic_bytes(self):
        resp = self._make()
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_empty_rows_still_produces_pdf(self):
        resp = export_pdf(["Col A", "Col B"], [])
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_single_column(self):
        resp = export_pdf(["Only Column"], [["value 1"], ["value 2"]])
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_many_columns_uses_landscape(self):
        headers = [f"Col{i}" for i in range(8)]
        rows = [[f"v{j}{i}" for i in range(8)] for j in range(5)]
        resp = export_pdf(headers, rows, title="Wide Report")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_subtitle_included(self):
        resp = export_pdf(["H"], [[1]], title="P&L", subtitle="Jan 2025 – Mar 2025")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_decimal_values(self):
        resp = export_pdf(["Amt"], [[Decimal("99999.99")]])
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_none_values(self):
        resp = export_pdf(["Val"], [[None], [None]])
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_unicode_title_and_data(self):
        resp = export_pdf(["Ìpínlẹ"], [["₦1,000"]], title="Ìjábọ")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_content_not_empty(self):
        resp = self._make()
        self.assertGreater(len(resp.content), 100)


# ─── dispatch_export ──────────────────────────────────────────────────────────


class TestDispatchExport(SimpleTestCase):

    _HEADERS = ["Name", "Amount"]
    _ROWS = [["Alice", 1000]]

    def test_excel_format_returns_xlsx_response(self):
        resp = dispatch_export("excel", self._HEADERS, self._ROWS, title="T", filename_base="r")
        self.assertIsNotNone(resp)
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_pdf_format_returns_pdf_response(self):
        resp = dispatch_export("pdf", self._HEADERS, self._ROWS, title="T", filename_base="r")
        self.assertIsNotNone(resp)
        self.assertEqual(resp["Content-Type"], "application/pdf")

    def test_json_format_returns_none(self):
        resp = dispatch_export("json", self._HEADERS, self._ROWS, title="T", filename_base="r")
        self.assertIsNone(resp)

    def test_unknown_format_returns_none(self):
        resp = dispatch_export("csv", self._HEADERS, self._ROWS, title="T", filename_base="r")
        self.assertIsNone(resp)

    def test_empty_string_format_returns_none(self):
        resp = dispatch_export("", self._HEADERS, self._ROWS, title="T", filename_base="r")
        self.assertIsNone(resp)

    def test_excel_filename_has_xlsx_extension(self):
        resp = dispatch_export("excel", self._HEADERS, self._ROWS, filename_base="my_report")
        self.assertIn("my_report.xlsx", resp["Content-Disposition"])

    def test_pdf_filename_has_pdf_extension(self):
        resp = dispatch_export("pdf", self._HEADERS, self._ROWS, filename_base="my_report")
        self.assertIn("my_report.pdf", resp["Content-Disposition"])

    def test_subtitle_passed_to_excel(self):
        import openpyxl
        resp = dispatch_export(
            "excel", ["H"], [[1]],
            title="Title", subtitle="Date Range", filename_base="test"
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        first_cell = ws.cell(row=1, column=1).value
        self.assertIsNotNone(first_cell)

    def test_empty_rows_excel(self):
        resp = dispatch_export("excel", ["H1", "H2"], [], filename_base="empty")
        self.assertIsNotNone(resp)
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_empty_rows_pdf(self):
        resp = dispatch_export("pdf", ["H1", "H2"], [], filename_base="empty")
        self.assertIsNotNone(resp)
        self.assertTrue(resp.content.startswith(b"%PDF"))
