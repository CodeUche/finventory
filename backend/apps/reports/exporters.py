"""
Report export utilities — Excel (.xlsx) and PDF.

Usage:
    from apps.reports.exporters import export_excel, export_pdf

    # In a view:
    return export_excel(headers, rows, sheet_name='P&L', filename='pnl.xlsx')
    return export_pdf(headers, rows, title='Profit & Loss', filename='pnl.pdf')

Both functions return a fully-formed Django HttpResponse ready to return from
a view.  They do NOT touch the database — callers are responsible for providing
clean headers (list[str]) and rows (list[list]).
"""

import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

from django.http import HttpResponse

# ─── Value normalisation ─────────────────────────────────────────────────────


def _clean(value: Any) -> Any:
    """
    Normalise a cell value to a type safe for both openpyxl and ReportLab.

    Rules
    -----
    * Decimal  → float  (Excel cells, ReportLab paragraphs both handle floats)
    * date / datetime → ISO string  (avoids timezone/tz-awareness issues)
    * None / '' → empty string
    * bool → 'Yes' / 'No'
    * Everything else → coerce to str only if not already a basic scalar
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    # UUIDs, enums, custom objects → string
    if not isinstance(value, (int, float, str)):
        return str(value)
    return value


# ─── Payload normalisation for export ────────────────────────────────────────
#
# The report registry (apps.reports.registry) has resolvers that return very
# different JSON shapes depending on what they report on: a flat list of rows,
# a dict with nested per-account/per-entry sections, a bare list, or a small
# dict of scalar summary figures. `export_excel`/`export_pdf` only understand
# one shape — `headers: list[str]` + `rows: list[list]` — so every resolver's
# output has to be normalised down to that before it can be exported.
#
# `flatten_for_export()` is that single normalisation point. It is shared by
# the single-report export path (ReportDispatchView in views.py) and the bulk
# multi-report export endpoint, so a report only needs to be taught how to
# flatten itself ONCE and both export paths get it "for free". It mirrors
# (and should be kept in sync with) the shape-aware rendering added to
# frontend/src/pages/reports/AllReportsPage.tsx — the same nested shapes are
# rendered on screen there and flattened to a table here.

SKIP_KEYS = {"period_start", "period_end", "as_of", "key", "label", "category"}


def _prettify(key: str) -> str:
    """'gross_sales' -> 'Gross Sales' — used to turn a dict key into a column/row label."""
    return str(key).replace("_", " ").title()


def _rows_from_dicts(items: "list[dict]"):
    """Flat list[dict] with a consistent key set -> (headers, rows), skipping
    any per-row keys that are internal bookkeeping (ids, drill-down refs)."""
    if not items:
        return [], []
    keys = [k for k in items[0].keys() if k not in ("id", "journal_entry_id")]
    headers = [_prettify(k) for k in keys]
    rows = [[item.get(k) for k in keys] for item in items]
    return headers, rows


def _kv_rows(pairs: "list[tuple[str, Any]]"):
    """A list of (label, value) pairs -> ('Item'/'Value' headers, rows) — used
    for reports that are fundamentally a handful of summary figures rather
    than a table (Profit & Loss, VAT/Tax Summary)."""
    return ["Item", "Value"], [[label, value] for label, value in pairs]


def flatten_for_export(data: Any):
    """
    Normalise any report resolver's return value into (headers, rows, totals)
    ready for export_excel()/export_pdf(), or return None if the shape isn't
    recognised (caller should then fall back to serving raw JSON).

    totals, when not None, is a dict of {column_label: value} to render as a
    bold grand-total row at the bottom of the sheet (see export_excel).
    """
    # 1) Bare list (e.g. trial_balance_report, which returns `list`, not `dict`).
    if isinstance(data, list):
        if not data or not isinstance(data[0], dict):
            return None
        headers, rows = _rows_from_dicts(data)
        # A trial balance's signed balances should always net to zero — surface
        # that as a sanity-check total rather than silently omitting a totals row.
        if all("balance" in d for d in data):
            total_balance = sum((_num(d.get("balance")) for d in data), Decimal("0"))
            return headers, rows, {"Balance": total_balance}
        return headers, rows, None

    if not isinstance(data, dict):
        return None

    # 2) Already-flat `rows` (the common case — most registry resolvers).
    if isinstance(data.get("rows"), list):
        headers, rows = _rows_from_dicts(data["rows"])
        totals = _totals_from_payload(data)
        return headers, rows, totals

    # 3) Key aliases that are already row-shaped lists under a different name.
    for alias in ("items", "groups"):
        if isinstance(data.get(alias), list):
            headers, rows = _rows_from_dicts(data[alias])
            totals = _totals_from_payload(data)
            return headers, rows, totals

    # 4) Per-account ledger detail (gl_detail / cash_register): one section per
    # account, each with an opening-balance row, its transaction lines, and a
    # closing-balance row — flattened into one table with the account repeated
    # per row (same pattern the sample Sage reports use for "Vendor Ledgers").
    if isinstance(data.get("accounts"), list):
        headers = ["Account Code", "Account Name", "Date", "Reference", "Description", "Debit", "Credit", "Balance"]
        rows: list = []
        for acct in data["accounts"]:
            code, name = acct.get("account_code", ""), acct.get("account_name", "")
            rows.append([code, name, "", "", "Opening Balance", "", "", acct.get("opening_balance")])
            for ln in acct.get("lines", []):
                rows.append([code, name, ln.get("date"), ln.get("reference"),
                             ln.get("description"), ln.get("debit"), ln.get("credit"), ln.get("balance")])
            rows.append([code, name, "", "", "Closing Balance", "", "", acct.get("closing_balance")])
        return headers, rows, None

    # 5) Journal register: one row per journal line, with the parent entry's
    # date/reference/description repeated on every line (same "repeat the
    # parent key" pattern used by the sample Purchase Journal/Sales Journal).
    if isinstance(data.get("entries"), list):
        headers = ["Date", "Reference", "Description", "Source Type",
                   "Account Code", "Account Name", "Line Description", "Debit", "Credit"]
        rows = []
        for entry in data["entries"]:
            for ln in entry.get("lines", []):
                rows.append([entry.get("date"), entry.get("reference"), entry.get("description"),
                             entry.get("source_type"), ln.get("account_code"), ln.get("account_name"),
                             ln.get("description"), ln.get("debit"), ln.get("credit")])
        totals = {}
        if "total_debit" in data:
            totals["Debit"] = data["total_debit"]
        if "total_credit" in data:
            totals["Credit"] = data["total_credit"]
        return headers, rows, (totals or None)

    # 6) Notes to the financial statements — a short numbered list.
    if isinstance(data.get("notes"), list):
        headers = ["Number", "Title", "Body"]
        rows = [[n.get("number"), n.get("title"), n.get("body")] for n in data["notes"]]
        return headers, rows, None

    # 7) Balance sheet: assets / liabilities / equity sections, each a list of
    # {code, name, balance} rows, plus grand totals for each section.
    if {"assets", "liabilities", "equity"} <= data.keys():
        headers = ["Section", "Code", "Name", "Balance"]
        rows = []
        for section, key in (("Assets", "assets"), ("Liabilities", "liabilities"), ("Equity", "equity")):
            for r in data.get(key, []):
                rows.append([section, r.get("code", ""), r.get("name", ""), r.get("balance")])
        totals = {
            "Total Assets": data.get("total_assets"),
            "Total Liabilities": data.get("total_liabilities"),
            "Total Equity": data.get("total_equity"),
        }
        return headers, rows, totals

    # 8) Profit & Loss: a handful of summary figures plus a nested `revenue`
    # breakdown — there's no natural "table" here, so export as key/value rows.
    if "revenue" in data and "gross_profit" in data:
        revenue = data.get("revenue") or {}
        pairs = [(f"Revenue — {_prettify(k)}", v) for k, v in revenue.items()]
        for k in ("cost_of_goods_sold", "gross_profit", "gross_margin_pct",
                  "operating_expenses", "miscellaneous_income", "net_profit", "net_margin_pct"):
            if k in data:
                pairs.append((_prettify(k), data[k]))
        headers, rows = _kv_rows(pairs)
        return headers, rows, None

    # 9) Tax Summary: VAT sub-dict plus WHT/PAYE scalars — same key/value treatment.
    if "vat" in data and ("wht_withheld" in data or "paye_payable" in data):
        vat = data.get("vat") or {}
        pairs = [(f"VAT — {_prettify(k)}", v) for k, v in vat.items() if k not in SKIP_KEYS]
        for k in ("wht_withheld", "paye_payable"):
            if k in data:
                pairs.append((_prettify(k), data[k]))
        headers, rows = _kv_rows(pairs)
        return headers, rows, None

    # 10) Financial Report Pack: P&L + Balance Sheet + Trial Balance bundled
    # together. Each sub-report has a different natural shape, so this is
    # deliberately reduced to one common "Section / Item / Value" table rather
    # than three different column layouts glued together.
    if {"profit_and_loss", "balance_sheet", "trial_balance"} <= data.keys():
        headers = ["Section", "Item", "Value"]
        rows = []
        pnl = data["profit_and_loss"]
        for k, v in (pnl.get("revenue") or {}).items():
            rows.append(["Profit & Loss", f"Revenue — {_prettify(k)}", v])
        for k in ("cost_of_goods_sold", "gross_profit", "operating_expenses", "net_profit"):
            if k in pnl:
                rows.append(["Profit & Loss", _prettify(k), pnl[k]])
        bs = data["balance_sheet"]
        for section, key in (("Assets", "assets"), ("Liabilities", "liabilities"), ("Equity", "equity")):
            for r in bs.get(key, []):
                rows.append(["Balance Sheet", f"{section} — {r.get('name', '')}", r.get("balance")])
        for k in ("total_assets", "total_liabilities", "total_equity"):
            if k in bs:
                rows.append(["Balance Sheet", _prettify(k), bs[k]])
        for r in data["trial_balance"]:
            rows.append(["Trial Balance", f"{r.get('code', '')} {r.get('name', '')}".strip(), r.get("balance")])
        return headers, rows, None

    # 11) Fallback: any other dict of mostly-scalar summary figures (e.g. Cash
    # Flow, Net Tax Report) — export the scalar keys as key/value rows rather
    # than giving up. Skips nested list/dict values we don't have a specific
    # renderer for, so partially-recognised payloads still export something
    # useful instead of nothing.
    scalar_pairs = [
        (_prettify(k), v) for k, v in data.items()
        if k not in SKIP_KEYS and not isinstance(v, (list, dict))
    ]
    if scalar_pairs:
        headers, rows = _kv_rows(scalar_pairs)
        return headers, rows, None

    return None


def _num(v) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _totals_from_payload(data: dict):
    """Pull a `total`/`totals` summary out of a `rows`-shaped payload, if present."""
    if isinstance(data.get("totals"), dict):
        return {_prettify(k): v for k, v in data["totals"].items()}
    if "total" in data and data["total"] is not None:
        return {"Total": data["total"]}
    return None


# ─── Excel export ─────────────────────────────────────────────────────────────

# Audity brand colour (dark navy)
_BRAND_HEX = "1E3A5F"
_ALT_ROW_HEX = "EEF2F7"
_WHITE = "FFFFFF"

# Same "does this column hold money" heuristic as the frontend's MONEY_HINT
# regex (frontend/src/pages/reports/AllReportsPage.tsx) — kept in sync so a
# column renders as currency identically on screen and in the exported file.
# Matched against the ORIGINAL (pre-prettified) snake_case header text.
_MONEY_HINT = re.compile(
    r"amount|total|balance|cost|revenue|debit|credit|net|gross|paye|salary|value"
    r"|outstanding|deduction|depreciation|subtotal|tax|pension|nhf",
    re.IGNORECASE,
)
_NOT_MONEY_HINT = re.compile(
    # "gross_margin_pct"/"net_margin_pct" contain "gross"/"net" so they'd
    # otherwise match _MONEY_HINT — exclude anything percent-flavoured.
    r"count|quantity|hours|assets|orders|level|pct|percent",
    re.IGNORECASE,
)


def _is_money_header(header: str) -> bool:
    return bool(_MONEY_HINT.search(header)) and not _NOT_MONEY_HINT.search(header)


# NOTE: _org_display_name() and _write_report_sheet() below are also imported
# directly by apps.reports.views.ReportBulkExportView (multi-sheet workbook
# builder) — despite the leading underscore they are shared, not file-private.
# Keep their signatures stable, or update both call sites together.

def _org_display_name(org) -> str:
    """Same org-name fallback chain export_pdf already uses, so Excel and PDF
    headers agree on what to call the organisation."""
    if org is None:
        return ""
    return (getattr(org, "invoice_company_name", None) or "").strip() or (getattr(org, "name", None) or "")


def _write_report_sheet(
    ws,
    headers: list[str],
    rows: list[list],
    *,
    title: str = "",
    subtitle: str = "",
    org_name: str = "",
    filter_note: str = "",
    totals: Optional[dict] = None,
) -> None:
    """
    Write one report's data into a worksheet, starting at row 1.

    Lays out a header block matching the reference sample reports (company
    name / report title / period / filter note, each its own row), then the
    branded column-header row, the data, and — if `totals` is supplied — a
    bold grand-total row at the bottom. Shared by export_excel() (single
    sheet) and the bulk multi-report export (one call per sheet), so both
    produce visually identical output.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color=_BRAND_HEX, end_color=_BRAND_HEX, fill_type="solid")
    header_font = Font(color=_WHITE, bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    alt_fill = PatternFill(start_color=_ALT_ROW_HEX, end_color=_ALT_ROW_HEX, fill_type="solid")
    total_border = Border(top=Side(style="thin", color="1E3A5F"))
    last_col = max(len(headers), 1)

    def _merged_text_row(row_num: int, text: str, font: Font):
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = font
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=last_col)

    # ── Header block: company name / report title / period / filter note ──────
    # Every row here is optional and simply omitted (not left blank) when the
    # caller doesn't have that piece of information for a given report, so a
    # report with no meaningful "filter" note doesn't get an empty row.
    row_num = 1
    if org_name:
        _merged_text_row(row_num, org_name, Font(bold=True, size=13, color="16161E"))
        row_num += 1
    if title:
        _merged_text_row(row_num, title, Font(bold=True, size=12, color="1E3A5F"))
        row_num += 1
    if subtitle:
        _merged_text_row(row_num, subtitle, Font(italic=True, color="666666", size=10))
        row_num += 1
    if filter_note:
        _merged_text_row(row_num, filter_note, Font(italic=True, color="888888", size=9))
        row_num += 1

    # ── Column header row ──────────────────────────────────────────────────────
    header_row = row_num
    money_col = [_is_money_header(h) for h in headers]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # ── Data rows ───────────────────────────────────────────────────────────────
    data_start = header_row + 1
    cleaned_rows = []
    for row_idx, row in enumerate(rows, data_start):
        cleaned = [_clean(v) for v in row]
        cleaned_rows.append(cleaned)
        is_alt = (row_idx - data_start) % 2 == 1
        for col_idx, value in enumerate(cleaned, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if money_col[col_idx - 1] and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = right
            else:
                cell.alignment = left
            if is_alt:
                cell.fill = alt_fill

    # ── Totals row ──────────────────────────────────────────────────────────────
    # `totals` maps a column label (matched case-insensitively against the
    # prettified header) to its summed value — most `rows`-shaped resolvers
    # already compute this (see registry.py), so it's usually a direct pass-
    # through of data already in the payload, not a recomputation here.
    total_row_idx = None
    if totals:
        total_row_idx = data_start + len(rows)
        lower_totals = {str(k).lower(): v for k, v in totals.items()}
        first_col_used = False
        for col_idx, header in enumerate(headers, 1):
            val = lower_totals.get(str(header).lower())
            if not first_col_used and val is None:
                # First column with no matching total value carries the "Total" label.
                ws.cell(row=total_row_idx, column=col_idx, value="Total")
                first_col_used = True
                continue
            if val is not None:
                cell = ws.cell(row=total_row_idx, column=col_idx, value=_clean(val))
                if money_col[col_idx - 1] and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                cell.alignment = right
                first_col_used = True
        for col_idx in range(1, last_col + 1):
            ws.cell(row=total_row_idx, column=col_idx).font = Font(bold=True)
            ws.cell(row=total_row_idx, column=col_idx).border = total_border

    # ── Column auto-width ───────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for cleaned in cleaned_rows:
            if col_idx - 1 < len(cleaned):
                max_len = max(max_len, len(str(cleaned[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Freeze header, header row height ────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.row_dimensions[header_row].height = 22


def export_excel(
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Report",
    filename: str = "report.xlsx",
    subtitle: str = "",
    title: str = "",
    org=None,
    totals: Optional[dict] = None,
    filter_note: str = "",
) -> HttpResponse:
    """
    Build a single-sheet .xlsx file and return it as an HttpResponse attachment.

    Args:
        headers:     Column header strings.
        rows:        Iterable of row iterables; values are normalised automatically.
        sheet_name:  Worksheet tab label (max 31 chars, truncated if necessary).
        filename:    Suggested download filename.
        subtitle:    Period / as-of label, e.g. "For the year 2025".
        title:       Report title, e.g. "Profit & Loss" — shown as its own header
                     row above the data (sheet_name is tab-length-limited and not
                     always a good on-page title, so this is separate).
        org:         Organisation instance — its name becomes the top header row,
                     same fallback chain export_pdf() uses.
        totals:      Optional {column_label: value} to render as a bold grand-total
                     row at the bottom (see flatten_for_export() in this module).
        filter_note: Optional one-line description of the report's scope/ordering,
                     matching the "Filter Criteria includes: ..." line the reference
                     sample reports show under the period. Omitted if not supplied.

    Returns:
        HttpResponse with content-type application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _write_report_sheet(
        ws, headers, rows,
        title=title, subtitle=subtitle, org_name=_org_display_name(org),
        filter_note=filter_note, totals=totals,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─── PDF export ───────────────────────────────────────────────────────────────


def export_pdf(
    headers: list[str],
    rows: list[list],
    title: str = "Report",
    subtitle: str = "",
    filename: str = "report.pdf",
    org=None,
) -> HttpResponse:
    """
    Build a PDF file and return it as an HttpResponse attachment.

    Layout (matches the client-side pdfUtils stock-card template):
      - Header every page: logo (top-left), company name/address/phone/email,
        document title (top-right), subtitle below title.
      - Rule line separating header from body.
      - Body: branded table with alternating row shading.
      - Footer every page: rule + "OrgName · Generated by Audity · Date"
        left, title centre, "Page X of Y" right.
      - Brand accent bar flush at bottom of last page only.
      - Landscape A4 when more than 6 columns; portrait A4 otherwise.

    Args:
        headers:  Column header strings.
        rows:     Iterable of row iterables; values are normalised automatically.
        title:    Document title (e.g. "Profit & Loss Statement").
        subtitle: Period label shown below the title (e.g. "Jan – Dec 2025").
        filename: Suggested download filename.
        org:      Organisation model instance (provides branding + contact info).

    Returns:
        HttpResponse with content-type application/pdf
    """
    import io as _io
    import urllib.request

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.pagesizes import landscape as rl_landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen.canvas import Canvas
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    # ── Org branding ──────────────────────────────────────────────────────────
    raw_brand = (getattr(org, "brand_color", None) or "").lstrip("#") or _BRAND_HEX
    brand_color = colors.HexColor(f"#{raw_brand}")

    org_name    = (
        (getattr(org, "invoice_company_name", None) or "").strip()
        or getattr(org, "name", None) or ""
    )
    org_address = getattr(org, "address", None) or ""
    org_phone   = getattr(org, "phone",   None) or ""
    org_email   = getattr(org, "email",   None) or ""
    logo_url    = getattr(org, "logo",    None) or ""

    # ── Fetch logo bytes (best-effort, 3 s timeout) ───────────────────────────
    logo_reader = None
    if logo_url:
        try:
            from reportlab.lib.utils import ImageReader
            req = urllib.request.Request(logo_url, headers={"User-Agent": "Audity/1.0"})
            with urllib.request.urlopen(req, timeout=3) as resp:
                logo_bytes = resp.read()
            logo_reader = ImageReader(_io.BytesIO(logo_bytes))
        except Exception:
            logo_reader = None

    # ── Page geometry ─────────────────────────────────────────────────────────
    is_landscape = len(headers) > 6
    page_size    = rl_landscape(A4) if is_landscape else A4
    page_w, page_h = page_size
    margin       = 15 * mm
    LOGO_SIZE    = 20 * mm
    HEADER_H     = 34 * mm   # top margin reserved for header block
    FOOTER_H     = 12 * mm   # bottom margin reserved for footer block

    today_str = datetime.today().strftime("%d %b %Y")

    # ── Numbered canvas — draws header + footer on every page ─────────────────
    class _HFCanvas(Canvas):
        def __init__(self, *args, **kwargs):
            Canvas.__init__(self, *args, **kwargs)
            self._pages: list[dict] = []

        def showPage(self):
            self._pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            n = len(self._pages)
            for idx, state in enumerate(self._pages):
                self.__dict__.update(state)
                self._draw(idx + 1, n)
                Canvas.showPage(self)
            Canvas.save(self)

        def _draw(self, page_num: int, total: int):
            self.saveState()
            pw, ph = self._pagesize

            # ── Header ───────────────────────────────────────────────────────
            logo_x = margin
            logo_y = ph - margin - LOGO_SIZE

            if logo_reader:
                try:
                    self.drawImage(
                        logo_reader, logo_x, logo_y,
                        LOGO_SIZE, LOGO_SIZE,
                        mask="auto", preserveAspectRatio=True,
                    )
                except Exception:
                    pass

            name_x = margin + (LOGO_SIZE + 4 * mm if logo_reader else 0)

            if org_name:
                self.setFont("Helvetica-Bold", 11)
                self.setFillColor(colors.HexColor("#16161E"))
                self.drawString(name_x, ph - margin - 8 * mm, org_name)

            self.setFont("Helvetica", 8)
            self.setFillColor(colors.HexColor("#6B7280"))
            info_y = ph - margin - 14 * mm
            for line in filter(None, [org_address, org_phone, org_email]):
                self.drawString(name_x, info_y, line)
                info_y -= 4.5 * mm

            # Document title — top-right, brand colour
            self.setFont("Helvetica-Bold", 16)
            self.setFillColor(brand_color)
            self.drawRightString(pw - margin, ph - margin - 8 * mm, title.upper())

            # Subtitle — below title
            if subtitle:
                self.setFont("Helvetica", 8)
                self.setFillColor(colors.HexColor("#6B7280"))
                self.drawRightString(pw - margin, ph - margin - 15 * mm, subtitle)

            # Rule below header
            rule_y = ph - HEADER_H
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.3)
            self.line(margin, rule_y, pw - margin, rule_y)

            # ── Footer ───────────────────────────────────────────────────────
            footer_rule_y = FOOTER_H - 2 * mm
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.25)
            self.line(margin, footer_rule_y, pw - margin, footer_rule_y)

            self.setFont("Helvetica", 6.5)
            self.setFillColor(colors.HexColor("#6B7280"))
            left_txt = f"{org_name}  ·  Generated by Audity  ·  {today_str}"
            txt_y = footer_rule_y - 3.5 * mm
            self.drawString(margin, txt_y, left_txt)
            self.drawCentredString(pw / 2, txt_y, title)
            self.drawRightString(pw - margin, txt_y, f"Page {page_num} of {total}")

            # Brand accent bar — last page only
            if page_num == total:
                self.setFillColor(brand_color)
                self.rect(0, 0, pw, 2 * mm, fill=1, stroke=0)

            self.restoreState()

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=HEADER_H + 4 * mm,
        bottomMargin=FOOTER_H + 4 * mm,
    )

    # ── Table ─────────────────────────────────────────────────────────────────
    alt_row_color = colors.HexColor(f"#{_ALT_ROW_HEX}")
    rule_color    = colors.HexColor("#E2E8F0")
    dark_color    = colors.HexColor("#16161E")

    table_data = [headers]
    for row in rows:
        table_data.append([str(_clean(v)) for v in row])

    usable_width = page_w - 2 * margin
    col_width    = usable_width / max(len(headers), 1)

    tbl = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)

    style_cmds: list = [
        ("BACKGROUND",    (0, 0), (-1,  0), brand_color),
        ("TEXTCOLOR",     (0, 0), (-1,  0), colors.white),
        ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,  0), 7.5),
        ("ALIGN",         (0, 0), (-1,  0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TEXTCOLOR",     (0, 1), (-1, -1), dark_color),
        ("ALIGN",         (0, 1), (-1, -1), "LEFT"),
        ("GRID",          (0, 0), (-1, -1), 0.2, rule_color),
        ("TOPPADDING",    (0, 0), (-1, -1), 2.8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3.5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3.5),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), alt_row_color))

    tbl.setStyle(TableStyle(style_cmds))

    elements: list = [tbl]

    doc.build(elements, canvasmaker=_HFCanvas)
    buf.seek(0)
    pdf_bytes = buf.getvalue()

    if org is not None:
        # maybe_save_pdf_to_drive never raises (see its own docstring) — a
        # Drive/Celery hiccup must never break a report download.
        from apps.connectors.services import maybe_save_pdf_to_drive
        maybe_save_pdf_to_drive(org, filename, pdf_bytes)

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─── Format dispatcher ────────────────────────────────────────────────────────


def dispatch_export(
    fmt: str,
    headers: list[str],
    rows: list[list],
    title: str = "Report",
    subtitle: str = "",
    filename_base: str = "report",
    org=None,
    totals: Optional[dict] = None,
) -> Optional[HttpResponse]:
    """
    Return an HttpResponse for Excel or PDF, or None if fmt is not an export format.

    Args:
        fmt:           'excel' | 'pdf' | anything else (returns None → caller serves JSON)
        headers:       Column headers.
        rows:          Data rows.
        title:         Report title — used as the Excel sheet name/title row and
                       the PDF document title.
        subtitle:      Sub-heading / date range string for both formats.
        filename_base: Base filename without extension (e.g. 'pnl_report').
        org:           Organisation instance — becomes the Excel company-name
                       header row and is forwarded to export_pdf for branding.
        totals:        Optional {column_label: value} grand-total row (Excel only
                       — export_pdf doesn't currently render a totals row).

    Returns:
        HttpResponse or None
    """
    if fmt == "excel":
        return export_excel(
            headers=headers,
            rows=rows,
            sheet_name=title[:31],
            filename=f"{filename_base}.xlsx",
            subtitle=subtitle,
            title=title,
            org=org,
            totals=totals,
        )
    if fmt == "pdf":
        return export_pdf(
            headers=headers,
            rows=rows,
            title=title,
            subtitle=subtitle,
            filename=f"{filename_base}.pdf",
            org=org,
        )
    return None
