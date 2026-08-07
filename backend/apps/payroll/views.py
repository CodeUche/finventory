import csv
import io
import json as _json
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter
import urllib.error
import urllib.parse
import urllib.request

from django.conf import settings as django_settings
from django.http import HttpResponse

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from apps.core.mixins import ExportMixin, TenantFilterMixin
from apps.core.permissions import IsManager, IsStaff, IsOwnerOrAdmin, plan_requires

_PlanPayroll = plan_requires('payroll')
from .models import (
    AdvancePolicy, AdvanceRequest, Attendance, BenefitPlan, Bonus, ClearanceChecklistItem,
    CompensationRecord, Employee, EmployeeBenefit, EmployeeDocument, EmployeeLoan,
    EmployeePenalty, EmployeeTaxProfile, ExitInterview, LeaveBalance, LeaveRequest,
    LeaveType, OffboardingCase, OffboardingChecklistTemplate, PayrollAdjustment,
    PayrollRun, PayrollSettings, PayslipDelivery, PayslipLine, PublicHoliday,
    StatutoryRemittance, TaxAuthority,
)
from .serializers import (
    AdvancePolicySerializer, AdvanceRequestSerializer, AttendanceSerializer,
    BenefitPlanSerializer, BonusSerializer, ClearanceChecklistItemSerializer,
    CompensationRecordSerializer, EmployeeBenefitSerializer, EmployeeDocumentSerializer,
    EmployeeLoanSerializer, EmployeePenaltySerializer, EmployeeSerializer,
    EmployeeTaxProfileSerializer, ExitInterviewSerializer, LeaveBalanceSerializer,
    LeaveRequestSerializer, LeaveTypeSerializer, OffboardingCaseSerializer,
    OffboardingChecklistTemplateSerializer, PayrollAdjustmentSerializer,
    PayrollRunSerializer, PayrollSettingsSerializer, PayslipDeliverySerializer,
    PayslipLineSerializer, PublicHolidaySerializer, StatutoryRemittanceSerializer,
    TaxAuthoritySerializer,
)
from .services import (
    CompensationService, EWAService, LeaveEncashmentService, LeaveService,
    OffboardingService, PayrollService, PublicHolidayService, RemittanceService,
    TaxAuthorityService, get_settings,
)


class EmployeeViewSet(ExportMixin, TenantFilterMixin, viewsets.ModelViewSet):
    export_filename = 'employees'
    export_fields = [
        ('Employee ID', 'employee_id'),
        ('First Name', 'first_name'),
        ('Last Name', 'last_name'),
        ('Email', 'email'),
        ('Phone', 'phone'),
        ('Job Title', 'job_title'),
        ('Department', 'department'),
        ('Type', 'employment_type'),
        ('Hire Date', 'hire_date'),
        ('Gross Salary', 'gross_salary'),
        ('Active', lambda o: 'Yes' if o.is_active else 'No'),
    ]
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]
    search_fields = ["first_name", "last_name", "email", "department", "job_title", "employee_id"]
    ordering_fields = ["first_name", "last_name", "basic_salary", "hire_date"]
    ordering = ["first_name"]

    def get_queryset(self):
        org = self._get_organisation()
        return Employee.objects.filter(organisation=org)

    def perform_create(self, serializer):
        # ATOMIC_REQUESTS is not enabled globally in this codebase, so wrap
        # explicitly: TenantFilterMixin.perform_create's save() and the
        # outbox emit() below must commit-or-rollback together (outbox
        # pattern guarantee — see apps.integrations.services).
        from django.db import transaction
        with transaction.atomic():
            super().perform_create(serializer)
            employee = serializer.instance
            from apps.integrations.services import IntegrationEventService
            IntegrationEventService.emit(
                employee.organisation, "employee.onboarded",
                {
                    "employee_id": str(employee.id),
                    "employee_code": employee.employee_id,
                    "full_name": f"{employee.first_name} {employee.last_name}".strip(),
                    "job_title": employee.job_title,
                    "department": employee.department,
                    "hire_date": str(employee.hire_date) if employee.hire_date else None,
                },
            )

    @action(detail=False, methods=["post"])
    def resolve_account(self, request):
        """POST /api/v1/payroll/employees/resolve_account/ — Resolve NUBAN account name.

        Tries Paystack then automatically falls back to Flutterwave (see
        apps.core.bank_resolve) so resolution doesn't depend on a single
        provider's account status.
        """
        from apps.core.bank_resolve import BankResolveError, resolve_account_name

        account_number = request.data.get("account_number", "").strip()
        bank_code = request.data.get("bank_code", "").strip()

        if not account_number or not bank_code:
            return Response({"error": "account_number and bank_code are required"}, status=400)

        try:
            account_name = resolve_account_name(account_number, bank_code)
        except BankResolveError as exc:
            return Response({"error": str(exc)}, status=400)
        return Response({"account_name": account_name})

    @action(detail=False, methods=['get'])
    def org_chart(self, request):
        """
        GET /payroll/employees/org_chart/

        Reporting tree built from the manager FK. Employees with no manager are
        roots; anyone caught in a cycle is surfaced as a root rather than
        dropped, so a bad edge is visible instead of silently hiding people.
        """
        org = self._get_organisation()
        employees = list(
            Employee.objects.filter(organisation=org, is_active=True)
            .values('id', 'employee_id', 'first_name', 'last_name',
                    'job_title', 'department', 'manager_id')
        )
        by_id = {str(e['id']): e for e in employees}
        for e in employees:
            e['id'] = str(e['id'])
            e['manager_id'] = str(e['manager_id']) if e['manager_id'] else None
            e['name'] = f"{e['first_name']} {e['last_name']}".strip()
            e['children'] = []

        roots = []
        for e in employees:
            parent_id = e['manager_id']
            # Walk up to detect a cycle before attaching.
            seen, cursor, cyclic = {e['id']}, parent_id, False
            while cursor:
                if cursor in seen:
                    cyclic = True
                    break
                seen.add(cursor)
                parent = by_id.get(cursor)
                cursor = str(parent['manager_id']) if parent and parent.get('manager_id') else None
            if parent_id and not cyclic and parent_id in by_id:
                by_id[parent_id]['children'].append(e)
            else:
                roots.append(e)
        return Response(roots)

    @action(detail=True, methods=['post'])
    def invite_portal(self, request, pk=None):
        """
        POST /payroll/employees/{id}/invite_portal/

        Creates a portal login for the employee so they can see their own
        payslips, leave and advances.

        The account is a normal email user, NOT an is_sub_account one: that flag
        forces the /staff-login route, which resolves a username against an
        organisation slug, and employees sign in with their own email address.
        What restricts them is the `employee` membership role, which carries no
        module permissions and is refused by every operator endpoint.
        """
        import secrets

        from apps.authentication.models import User
        from apps.tenancy.models import Membership

        employee = self.get_object()
        org = self._get_organisation()

        if employee.user_id:
            return Response({'error': 'This employee already has portal access.'}, status=400)
        email = (employee.email or '').strip().lower()
        if not email:
            return Response(
                {'error': 'Add an email address to this employee before inviting them.'},
                status=400,
            )
        if User.objects.filter(email__iexact=email).exists():
            return Response(
                {'error': 'A user with that email address already exists.'}, status=400,
            )

        temp_password = secrets.token_urlsafe(9)
        user = User.objects.create_user(
            email=email,
            password=temp_password,
            first_name=employee.first_name,
            last_name=employee.last_name,
            is_verified=True,
        )
        user.must_change_password = True
        user.save(update_fields=['must_change_password'])

        Membership.objects.get_or_create(
            organisation=org, user=user,
            defaults={'role': Membership.Role.EMPLOYEE, 'is_active': True},
        )
        employee.user = user
        employee.save(update_fields=['user'])

        try:
            from django.core.mail import send_mail
            send_mail(
                subject=f"Your {org.name} employee portal account",
                message=(
                    f"Hello {employee.first_name},\n\n"
                    f"An employee portal account has been created for you.\n\n"
                    f"Email: {email}\n"
                    f"Temporary password: {temp_password}\n\n"
                    f"You will be asked to change it when you first sign in.\n\n"
                    f"{org.name}"
                ),
                from_email=None,
                recipient_list=[email],
                fail_silently=True,
            )
            emailed = True
        except Exception:
            emailed = False

        return Response({
            'employee': str(employee.id),
            'email': email,
            'emailed': emailed,
            # Returned so an admin can pass it on when SMTP is not configured.
            'temporary_password': temp_password,
        }, status=201)

    @action(detail=True, methods=['post'])
    def revoke_portal(self, request, pk=None):
        """POST /payroll/employees/{id}/revoke_portal/ — disable portal access."""
        from apps.tenancy.models import Membership

        employee = self.get_object()
        if not employee.user_id:
            return Response({'error': 'This employee does not have portal access.'}, status=400)
        user = employee.user
        Membership.objects.filter(organisation=employee.organisation, user=user).update(
            is_active=False
        )
        user.is_active = False
        user.save(update_fields=['is_active'])
        employee.user = None
        employee.save(update_fields=['user'])
        return Response({'employee': str(employee.id), 'revoked': True})

    @action(detail=False, methods=['get'])
    def lifecycle_alerts(self, request):
        """
        GET /payroll/employees/lifecycle_alerts/?within_days=30

        Aggregates three kinds of live-computed alerts from existing Employee
        fields — no new models: probation ending soon (confirmation_date),
        contract ending soon (contract_end_date + employment_type=CONTRACT),
        and work anniversaries (hire_date). One query per kind.
        """
        from datetime import timedelta as _td

        from django.utils import timezone

        org = self._get_organisation()
        try:
            within_days = int(request.query_params.get('within_days', 30))
        except (TypeError, ValueError):
            within_days = 30
        today = timezone.localdate()
        horizon = today + _td(days=within_days)

        probation_ending = list(
            Employee.objects.filter(
                organisation=org, is_active=True,
                confirmation_date__gte=today, confirmation_date__lte=horizon,
            ).values('id', 'employee_id', 'first_name', 'last_name', 'confirmation_date', 'department')
        )
        contract_ending = list(
            Employee.objects.filter(
                organisation=org, is_active=True, employment_type=Employee.CONTRACT,
                contract_end_date__gte=today, contract_end_date__lte=horizon,
            ).values('id', 'employee_id', 'first_name', 'last_name', 'contract_end_date', 'department')
        )

        anniversaries = []
        for emp in Employee.objects.filter(
            organisation=org, is_active=True, hire_date__isnull=False,
        ).values('id', 'employee_id', 'first_name', 'last_name', 'hire_date', 'department'):
            hire = emp['hire_date']
            this_year_anniv = hire.replace(year=today.year)
            if this_year_anniv < today:
                this_year_anniv = hire.replace(year=today.year + 1)
            if today <= this_year_anniv <= horizon:
                years = this_year_anniv.year - hire.year
                anniversaries.append({**emp, 'anniversary_date': this_year_anniv, 'years': years})

        def _fmt(rows, date_key):
            for r in rows:
                r['id'] = str(r['id'])
                r['name'] = f"{r['first_name']} {r['last_name']}".strip()
            return sorted(rows, key=lambda r: r[date_key])

        return Response({
            'probation_ending': _fmt(probation_ending, 'confirmation_date'),
            'contract_ending': _fmt(contract_ending, 'contract_end_date'),
            'work_anniversaries': _fmt(anniversaries, 'anniversary_date'),
        })


class EmployeeDocumentViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = EmployeeDocumentSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = EmployeeDocument.objects.filter(organisation=org)
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs

    def perform_create(self, serializer):
        file = self.request.FILES.get('file')
        size = file.size if file else 0
        serializer.save(organisation=self._get_organisation(), file_size=size)

    @action(detail=False, methods=['get'])
    def expiring(self, request):
        """GET /payroll/documents/expiring/?within_days=30 — soonest-first."""
        from datetime import timedelta

        from django.utils import timezone

        org = self._get_organisation()
        try:
            within_days = int(request.query_params.get('within_days', 30))
        except (TypeError, ValueError):
            within_days = 30
        today = timezone.localdate()
        horizon = today + timedelta(days=within_days)
        qs = (
            EmployeeDocument.objects.filter(
                organisation=org, expiry_date__isnull=False,
                expiry_date__gte=today, expiry_date__lte=horizon,
            )
            .select_related('employee')
            .order_by('expiry_date')
        )
        return Response(EmployeeDocumentSerializer(qs, many=True, context={'request': request}).data)


class EmployeePenaltyViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = EmployeePenaltySerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = EmployeePenalty.objects.filter(organisation=org)
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=True, methods=['post'])
    def waive(self, request, pk=None):
        penalty = self.get_object()
        if penalty.status != EmployeePenalty.PENDING:
            return Response({'error': 'Only pending penalties can be waived'}, status=400)
        penalty.status = EmployeePenalty.WAIVED
        penalty.save(update_fields=['status'])
        return Response(EmployeePenaltySerializer(penalty).data)


class EmployeeLoanViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = EmployeeLoanSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = EmployeeLoan.objects.filter(organisation=org)
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        loan = self.get_object()
        if loan.status != EmployeeLoan.ACTIVE:
            return Response({'error': 'Only active loans can be cancelled'}, status=400)
        loan.status = EmployeeLoan.CANCELLED
        loan.save(update_fields=['status'])
        return Response(EmployeeLoanSerializer(loan).data)


class BonusViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = BonusSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = Bonus.objects.filter(organisation=org).select_related('employee')
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        period_year = self.request.query_params.get('period_year')
        period_month = self.request.query_params.get('period_month')
        if period_year:
            qs = qs.filter(period_year=period_year)
        if period_month:
            qs = qs.filter(period_month=period_month)
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


class AttendanceViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = Attendance.objects.filter(organisation=org).select_related('employee')
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if year:
            qs = qs.filter(date__year=year)
        if month:
            qs = qs.filter(date__month=month)
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=False, methods=['post'])
    def bulk_mark(self, request):
        """
        POST /payroll/attendance/bulk_mark/
        Body: { employee_ids: [...], date: "YYYY-MM-DD", status: "present"|"absent"|..., overtime_hours: 0 }
        Marks attendance for multiple employees on the same day.
        """
        org = self._get_organisation()
        emp_ids = request.data.get('employee_ids', [])
        date = request.data.get('date')
        att_status = request.data.get('status', Attendance.PRESENT)
        overtime_hours = request.data.get('overtime_hours', 0)
        notes = request.data.get('notes', '')

        if not emp_ids or not date:
            return Response({'error': 'employee_ids and date are required'}, status=400)

        # Validate every submitted ID belongs to this organisation (prevents IDOR)
        valid_ids = set(
            Employee.objects.filter(organisation=org, id__in=emp_ids)
            .values_list('id', flat=True)
        )
        invalid = [str(e) for e in emp_ids if str(e) not in {str(v) for v in valid_ids}]
        if invalid:
            return Response(
                {'error': 'One or more employees not found in your organisation.'},
                status=403,
            )

        created, updated = 0, 0
        for emp_id in valid_ids:
            obj, is_new = Attendance.objects.update_or_create(
                organisation=org,
                employee_id=emp_id,
                date=date,
                defaults={
                    'status': att_status,
                    'overtime_hours': overtime_hours,
                    'notes': notes,
                }
            )
            if is_new:
                created += 1
            else:
                updated += 1

        return Response({'created': created, 'updated': updated})


class PayrollRunViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = PayrollRunSerializer
    permission_classes = [IsAuthenticated, IsManager, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        return PayrollRun.objects.filter(organisation=org).prefetch_related('payslips__employee')

    def create(self, request, *args, **kwargs):
        import calendar as _cal
        from datetime import date as _date

        org = self._get_organisation()
        try:
            year = int(request.data.get('period_year'))
            month = int(request.data.get('period_month'))
        except (TypeError, ValueError):
            return Response({'error': 'period_year and period_month are required.'}, status=400)
        if not 1 <= month <= 12:
            return Response({'error': 'period_month must be between 1 and 12.'}, status=400)

        run_type = request.data.get('run_type') or PayrollRun.REGULAR
        valid_types = [c for c, _ in PayrollRun.RUN_TYPE_CHOICES]
        if run_type not in valid_types:
            return Response(
                {'error': f"run_type must be one of: {', '.join(valid_types)}"}, status=400
            )

        # Optional explicit period window (drives proration for part-month runs)
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        if not period_start:
            period_start = _date(year, month, 1)
        if not period_end:
            period_end = _date(year, month, _cal.monthrange(year, month)[1])

        # Period-lock guard
        from apps.accounting.services import AccountingService
        if AccountingService.is_period_locked(org, _date(year, month, 1), user=request.user):
            return Response(
                {'error': f'The period {year}-{month:02d} is locked. Unlock it before running payroll.'},
                status=403,
            )

        if run_type == PayrollRun.REGULAR:
            # The regular run stays one-per-month: re-posting it re-runs in
            # place rather than creating a duplicate set of payslips.
            run, created = PayrollRun.objects.get_or_create(
                organisation=org, period_year=year, period_month=month,
                run_type=PayrollRun.REGULAR, sequence=1,
                defaults={
                    'processed_by': request.user,
                    'period_start': period_start,
                    'period_end': period_end,
                    'pay_frequency': request.data.get('pay_frequency')
                    or get_settings(org).default_pay_frequency,
                },
            )
            if not created and run.status != PayrollRun.DRAFT:
                # Re-running is destructive: it deletes and rebuilds payslips,
                # losing transfer references, and bonuses already marked APPLIED
                # would not be picked up a second time. Force the caller to be
                # explicit via recalculate/, or to raise an off-cycle run.
                return Response(
                    {'error': f'Payroll for {year}-{month:02d} has already been processed. '
                              f'Use recalculate to rebuild it, or create an off-cycle run.'},
                    status=400,
                )
        else:
            # Off-cycle / supplementary / 13th-month / final settlement runs
            # stack behind the regular run, each with its own sequence.
            last = (
                PayrollRun.objects
                .filter(organisation=org, period_year=year, period_month=month, run_type=run_type)
                .order_by('-sequence').first()
            )
            next_seq = (last.sequence + 1) if last else 1
            run = PayrollRun.objects.create(
                organisation=org, period_year=year, period_month=month,
                run_type=run_type, sequence=next_seq,
                processed_by=request.user,
                period_start=period_start, period_end=period_end,
                pay_frequency=request.data.get('pay_frequency')
                or get_settings(org).default_pay_frequency,
            )
            created = True

        run = PayrollService.run_payroll(run)
        return Response(
            PayrollRunSerializer(run).data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'])
    def recalculate(self, request, pk=None):
        """
        POST /payroll/runs/{id}/recalculate/

        Rebuild a processed run in place. Destructive by design — payslips are
        deleted and recreated, so transfer references are lost — and therefore
        refused once the run has been approved or paid.
        """
        run = self.get_object()
        if run.status in [PayrollRun.APPROVED, PayrollRun.PAID]:
            return Response(
                {'error': 'An approved or paid run cannot be recalculated. '
                          'Raise a supplementary run instead.'},
                status=400,
            )
        if run.payslips.filter(
            transfer_status__in=[PayslipLine.TRANSFER_INITIATED, PayslipLine.TRANSFER_SUCCESS]
        ).exists():
            return Response(
                {'error': 'Some salaries have already been transferred for this run. '
                          'Recalculating would lose those transfer references.'},
                status=400,
            )
        run = PayrollService.run_payroll(run)
        return Response(PayrollRunSerializer(run).data)

    @action(detail=False, methods=['get'])
    def eligible_approvers(self, request):
        """Return org members with admin or owner role — valid approvers for payroll."""
        from apps.tenancy.models import Membership
        org = self._get_organisation()
        members = (
            Membership.objects
            .filter(organisation=org, is_active=True, role__in=['owner', 'admin'])
            .select_related('user')
            .exclude(user=request.user)
        )
        data = [
            {'id': str(m.user.id), 'name': f"{m.user.first_name} {m.user.last_name}".strip() or m.user.email, 'email': m.user.email, 'role': m.role}
            for m in members
        ]
        return Response(data)

    @action(detail=True, methods=['post'])
    def submit_for_approval(self, request, pk=None):
        """Manager/HR submits a processed payroll for admin/owner approval."""
        run = self.get_object()
        if run.status != PayrollRun.PROCESSING:
            return Response({'error': 'Only processing payrolls can be submitted for approval'}, status=400)
        run.submitted_for_approval = True
        run.submitted_by = request.user

        # Optional: target a specific approver — must be an active admin/owner in this org
        approver_id = request.data.get('approver_id')
        if approver_id:
            from apps.tenancy.models import Membership
            try:
                membership = Membership.objects.select_related('user').get(
                    organisation=run.organisation,
                    user__id=approver_id,
                    is_active=True,
                    role__in=['admin', 'owner'],
                )
                run.target_approver = membership.user
            except Membership.DoesNotExist:
                return Response(
                    {'error': 'Approver must be an active admin or owner of this organisation.'},
                    status=400,
                )

        run.save(update_fields=['submitted_for_approval', 'submitted_by', 'target_approver'])

        # Notify admins/owners via audit log (notifications picked up by frontend polling)
        try:
            from apps.core.utils import log_audit
            log_audit(request, run, 'PAYROLL_SUBMITTED',
                      f"Payroll {run.run_number} submitted for approval by {request.user.email}")
        except Exception:
            pass

        return Response(PayrollRunSerializer(run).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Owner/Admin final approval — requires submitted_for_approval=True."""
        import logging as _log
        run = self.get_object()
        if run.status != PayrollRun.PROCESSING:
            return Response({'error': 'Only processing payrolls can be approved'}, status=400)
        if not run.submitted_for_approval:
            from apps.core.permissions import IsOwnerOrAdmin as _IsOwner
            checker = _IsOwner()
            if not checker.has_permission(request, self):
                return Response(
                    {'error': 'This payroll must be submitted for approval first.'},
                    status=400,
                )
        # Segregation of duties: the person who submitted cannot approve
        if run.submitted_by and run.submitted_by == request.user:
            return Response(
                {'error': 'You cannot approve a payroll run you submitted. A different owner or admin must approve.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        run.status = PayrollRun.APPROVED
        run.approved_by = request.user
        run.save()

        # Auto-post payroll journal entry (non-blocking)
        from apps.accounting.services import AccountingService, safe_post_gl
        safe_post_gl(
            AccountingService.post_payroll_journal, run.organisation, run, request.user,
            model_instance=run,
        )

        return Response(PayrollRunSerializer(run).data)

    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        from django.db import transaction as _tx
        from django.utils import timezone
        with _tx.atomic():
            run = PayrollRun.objects.select_for_update().get(
                pk=self.get_object().pk,
                organisation=self._get_organisation(),
            )
            if run.status != PayrollRun.APPROVED:
                return Response(
                    {'error': 'Only approved payrolls can be marked as paid.'},
                    status=400,
                )
            run.status = PayrollRun.PAID
            run.payment_date = request.data.get('payment_date', timezone.now().date())
            run.save()
            run.payslips.filter(status=PayslipLine.CALCULATED).update(status=PayslipLine.PAID)
        return Response(PayrollRunSerializer(run).data)

    @action(detail=True, methods=['post'])
    def initiate_transfers(self, request, pk=None):
        """
        POST /payroll/runs/{id}/initiate_transfers/
        Bulk Paystack salary transfer. Persists per-payslip transfer_status for reconciliation.
        """
        import logging as _log
        logger = _log.getLogger(__name__)

        run = self.get_object()
        if run.status != PayrollRun.APPROVED:
            return Response({'error': 'Only approved payroll runs can initiate transfers'}, status=400)

        secret_key = getattr(django_settings, 'PAYSTACK_SECRET_KEY', '')
        if not secret_key:
            return Response(
                {'error': 'Paystack is not configured. Add PAYSTACK_SECRET_KEY to your .env file.'},
                status=503,
            )

        def paystack_post(path, payload):
            req_body = _json.dumps(payload).encode()
            req = urllib.request.Request(
                f'https://api.paystack.co{path}',
                data=req_body,
                headers={
                    'Authorization': f'Bearer {secret_key}',
                    'Content-Type': 'application/json',
                    # Without a User-Agent, Paystack's Cloudflare front-end blocks
                    # requests from datacenter IPs (Railway) with a 403 "error code:
                    # 1010" before they ever reach Paystack's own API.
                    'User-Agent': 'Mozilla/5.0 (compatible; AudityBackend/1.0)',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                return _json.loads(resp.read())

        results = []
        transfers = []
        payslip_map = {}  # reference → payslip id

        for payslip in run.payslips.select_related('employee').all():
            emp = payslip.employee
            net = float(payslip.net_salary)
            ref = f'{run.run_number}-{emp.employee_id}'

            if not emp.account_number or not emp.bank_code:
                payslip.transfer_status = PayslipLine.TRANSFER_SKIPPED
                payslip.transfer_error = 'Missing bank account number or bank code'
                payslip.save(update_fields=['transfer_status', 'transfer_error'])
                results.append({
                    'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                    'status': 'skipped', 'reason': 'Missing bank account number or bank code',
                })
                continue

            # Create/get Paystack recipient code
            recipient_code = emp.paystack_recipient_code
            if not recipient_code:
                try:
                    r = paystack_post('/transferrecipient', {
                        'type': 'nuban',
                        'name': emp.account_name or f'{emp.first_name} {emp.last_name}',
                        'account_number': emp.account_number,
                        'bank_code': emp.bank_code,
                        'currency': 'NGN',
                    })
                    recipient_code = r['data']['recipient_code']
                    emp.paystack_recipient_code = recipient_code
                    emp.save(update_fields=['paystack_recipient_code'])
                except Exception as e:
                    logger.warning('Paystack create recipient failed for %s: %s', emp.employee_id, e)
                    payslip.transfer_status = PayslipLine.TRANSFER_FAILED
                    payslip.transfer_error = 'Could not create transfer recipient'
                    payslip.save(update_fields=['transfer_status', 'transfer_error'])
                    results.append({
                        'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                        'status': 'failed', 'reason': 'Could not create transfer recipient',
                    })
                    continue

            transfers.append({
                'amount': int(net * 100),
                'recipient': recipient_code,
                'reason': f'Net pay — {run.run_number}',
                'reference': ref,
            })
            payslip_map[ref] = payslip.id
            results.append({
                'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                'account': emp.account_number, 'bank': emp.bank_name,
                'amount': net, 'recipient_code': recipient_code,
                'status': 'queued', 'reference': ref,
            })

        if not transfers:
            return Response({
                'success': False,
                'message': 'No employees with complete bank details found',
                'results': results,
            }, status=400)

        try:
            bulk_resp = paystack_post('/transfer/bulk', {
                'currency': 'NGN', 'source': 'balance', 'transfers': transfers,
            })
            batch_code = (
                bulk_resp.get('data', {}).get('batch_code', '')
                if isinstance(bulk_resp.get('data'), dict)
                else ''
            )
            if not batch_code and isinstance(bulk_resp.get('data'), list):
                batch_code = ','.join(str(t.get('reference', '')) for t in bulk_resp['data'][:3])

            run.transfer_reference = batch_code or run.run_number
            run.save(update_fields=['transfer_reference'])

            # Persist per-payslip transfer status
            if isinstance(bulk_resp.get('data'), list):
                for transfer_result in bulk_resp['data']:
                    ref = transfer_result.get('reference', '')
                    pslip_id = payslip_map.get(ref)
                    if pslip_id:
                        t_status = transfer_result.get('status', '')
                        ps_status = (
                            PayslipLine.TRANSFER_INITIATED if t_status in ('pending', 'otp')
                            else PayslipLine.TRANSFER_SUCCESS if t_status == 'success'
                            else PayslipLine.TRANSFER_FAILED
                        )
                        PayslipLine.objects.filter(id=pslip_id).update(
                            transfer_status=ps_status,
                            transfer_reference=transfer_result.get('transfer_code', ref),
                        )
                        # Mirror to results list
                        for r in results:
                            if r.get('reference') == ref:
                                r['status'] = 'initiated'
                                r['transfer_code'] = transfer_result.get('transfer_code', '')
            else:
                # Fallback: mark all queued as initiated
                PayslipLine.objects.filter(
                    payroll_run=run,
                    transfer_status=PayslipLine.TRANSFER_PENDING,
                ).update(transfer_status=PayslipLine.TRANSFER_INITIATED)
                for r in results:
                    if r['status'] == 'queued':
                        r['status'] = 'initiated'

            return Response({
                'success': True,
                'message': f'{len(transfers)} transfer(s) initiated via Paystack',
                'batch_code': batch_code,
                'results': results,
            })

        except urllib.error.HTTPError as e:
            body = e.read().decode()
            logger.warning('Paystack bulk transfer failed: %s %s', e.code, body)
            try:
                err_data = _json.loads(body)
                msg = err_data.get('message', 'Paystack rejected the bulk transfer')
            except Exception:
                msg = f'HTTP {e.code}: {body[:200]}'
            return Response({'success': False, 'error': msg, 'results': results}, status=400)
        except Exception as e:
            logger.warning('Paystack bulk transfer error: %s', e)
            return Response({
                'success': False, 'error': 'Transfer service temporarily unavailable',
                'results': results,
            }, status=503)

    @action(detail=True, methods=['post'])
    def retry_failed(self, request, pk=None):
        """
        POST /payroll/runs/{id}/retry_failed/
        Retries only the payslips whose transfer_status is 'failed'.
        Same idempotency key (run_number-employee_id) prevents double payments.
        """
        import logging as _log
        logger = _log.getLogger(__name__)

        run = self.get_object()
        if run.status not in [PayrollRun.APPROVED, PayrollRun.PAID]:
            return Response({'error': 'Can only retry on approved or paid payroll runs'}, status=400)

        secret_key = getattr(django_settings, 'PAYSTACK_SECRET_KEY', '')
        if not secret_key:
            return Response({'error': 'Paystack is not configured.'}, status=503)

        failed_payslips = list(
            run.payslips.select_related('employee').filter(
                transfer_status__in=[PayslipLine.TRANSFER_FAILED, PayslipLine.TRANSFER_PENDING]
            )
        )
        if not failed_payslips:
            return Response({'message': 'No failed transfers to retry', 'retried': 0})

        def paystack_post(path, payload):
            req_body = _json.dumps(payload).encode()
            req = urllib.request.Request(
                f'https://api.paystack.co{path}',
                data=req_body,
                headers={
                    'Authorization': f'Bearer {secret_key}',
                    'Content-Type': 'application/json',
                    # Without a User-Agent, Paystack's Cloudflare front-end blocks
                    # requests from datacenter IPs (Railway) with a 403 "error code:
                    # 1010" before they ever reach Paystack's own API.
                    'User-Agent': 'Mozilla/5.0 (compatible; AudityBackend/1.0)',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                return _json.loads(resp.read())

        results = []
        transfers = []
        payslip_map = {}

        for payslip in failed_payslips:
            emp = payslip.employee
            if not emp.account_number or not emp.bank_code:
                results.append({'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                                 'status': 'skipped', 'reason': 'No bank details'})
                continue
            recipient_code = emp.paystack_recipient_code
            if not recipient_code:
                try:
                    r = paystack_post('/transferrecipient', {
                        'type': 'nuban',
                        'name': emp.account_name or f'{emp.first_name} {emp.last_name}',
                        'account_number': emp.account_number,
                        'bank_code': emp.bank_code,
                        'currency': 'NGN',
                    })
                    recipient_code = r['data']['recipient_code']
                    emp.paystack_recipient_code = recipient_code
                    emp.save(update_fields=['paystack_recipient_code'])
                except Exception as e:
                    logger.warning('Retry: recipient failed for %s: %s', emp.employee_id, e)
                    results.append({'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                                     'status': 'failed', 'reason': 'Could not create recipient'})
                    continue

            ref = f'{run.run_number}-{emp.employee_id}'
            transfers.append({
                'amount': int(float(payslip.net_salary) * 100),
                'recipient': recipient_code,
                'reason': f'Retry net pay — {run.run_number}',
                'reference': ref,
            })
            payslip_map[ref] = payslip.id
            results.append({'employee': emp.employee_id, 'name': f'{emp.first_name} {emp.last_name}',
                             'status': 'queued', 'reference': ref})

        if not transfers:
            return Response({'success': False, 'message': 'No retryable transfers', 'results': results})

        try:
            bulk_resp = paystack_post('/transfer/bulk', {
                'currency': 'NGN', 'source': 'balance', 'transfers': transfers,
            })
            for transfer_result in (bulk_resp.get('data') or []):
                ref = transfer_result.get('reference', '')
                pslip_id = payslip_map.get(ref)
                if pslip_id:
                    PayslipLine.objects.filter(id=pslip_id).update(
                        transfer_status=PayslipLine.TRANSFER_INITIATED,
                        transfer_reference=transfer_result.get('transfer_code', ref),
                        transfer_error='',
                    )
                    for r in results:
                        if r.get('reference') == ref:
                            r['status'] = 'initiated'

            return Response({'success': True, 'retried': len(transfers), 'results': results})
        except Exception as e:
            logger.warning('Retry bulk transfer error: %s', e)
            return Response({'success': False, 'error': 'Transfer retry failed. Please try again.', 'results': results}, status=503)

    @action(detail=True, methods=['get'])
    def export_bank_file(self, request, pk=None):
        """
        GET /payroll/runs/{id}/export_bank_file/

        Returns a professionally formatted Excel workbook (.xlsx) for bank
        submission — NIBSS EFT / NIP compatible.

        Workbook structure
        ------------------
        Sheet 1 — Payment Schedule   (ready-to-transfer employees, full breakdown)
        Sheet 2 — Statutory Summary  (employer remittance obligations)
        Sheet 3 — Exceptions         (employees with missing bank details)
        """
        from datetime import date as _date

        run  = self.get_object()
        org  = run.organisation
        curr = org.currency or 'NGN'
        payslips = run.payslips.select_related('employee').order_by(
            'employee__department', 'employee__last_name'
        )

        MONTHS = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']

        period_label  = f"{MONTHS[run.period_month]} {run.period_year}"
        generated_on  = _date.today().strftime('%d %B %Y')
        payment_date  = (run.payment_date.strftime('%d %B %Y')
                         if run.payment_date else 'Pending')
        processed_by  = (
            f"{run.processed_by.first_name} {run.processed_by.last_name}".strip()
            or run.processed_by.email
        )
        approved_by = (
            (f"{run.approved_by.first_name} {run.approved_by.last_name}".strip()
             or run.approved_by.email)
            if run.approved_by else 'Pending Approval'
        )

        def _d(v):
            try:
                return float(v or 0)
            except Exception:
                return 0.0

        ready, exceptions = [], []
        for p in payslips:
            e = p.employee
            if (e.account_number or '').strip() and (e.bank_code or '').strip():
                ready.append(p)
            else:
                exceptions.append(p)

        # Aggregate totals (ready employees only)
        tot_basic     = sum(_d(p.basic_salary)          for p in ready)
        tot_housing   = sum(_d(p.housing_allowance)     for p in ready)
        tot_transport = sum(_d(p.transport_allowance)   for p in ready)
        tot_leave     = sum(_d(p.leave_allowance)       for p in ready)
        tot_other     = sum(_d(p.other_allowances)      for p in ready)
        tot_gross     = sum(_d(p.gross_salary)          for p in ready)
        tot_bonus     = sum(_d(p.bonus_amount)          for p in ready)
        tot_overtime  = sum(_d(p.overtime_amount)       for p in ready)
        tot_earnings  = tot_gross + tot_bonus + tot_overtime
        tot_pen_emp   = sum(_d(p.employee_pension)      for p in ready)
        tot_nhf       = sum(_d(p.nhf)                   for p in ready)
        tot_nsitf     = sum(_d(p.nsitf)                 for p in ready)
        tot_paye      = sum(_d(p.paye_tax)              for p in ready)
        tot_att       = sum(_d(p.attendance_deduction)  for p in ready)
        tot_loan      = sum(_d(p.loan_deductions)       for p in ready)
        tot_penalty   = sum(_d(p.penalty_deductions)    for p in ready)
        tot_deduct    = sum(_d(p.total_deductions)      for p in ready)
        tot_net       = sum(_d(p.net_salary)            for p in ready)
        tot_pen_er    = _d(run.total_pension_employer)

        # ── Shared style helpers ─────────────────────────────────────────────────

        # Palette (navy + gold — investment-grade standard)
        NAVY   = '0D1F3C'   # header background
        GOLD   = 'C9A84C'   # accent / totals row
        WHITE  = 'FFFFFF'
        LIGHT  = 'F4F6FA'   # alternate row tint
        MID    = 'DDE3ED'   # section header tint
        RED_BG = 'FFF0F0'   # exception rows
        RED_TX = 'C0392B'

        def _fill(hex_color):
            return PatternFill('solid', fgColor=hex_color)

        def _font(bold=False, color=WHITE, size=10, italic=False):
            return Font(name='Calibri', bold=bold, color=color, size=size,
                        italic=italic)

        def _border(style='thin'):
            s = Side(style=style, color='B0BAC9')
            return Border(left=s, right=s, top=s, bottom=s)

        def _align(h='left', v='center', wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        MONEY_FMT  = f'"{curr}" #,##0.00'
        NUMBER_FMT = '#,##0'

        def _set_col_widths(ws, widths):
            for col_idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = w

        def _freeze(ws, cell='A1'):
            ws.freeze_panes = cell

        # ── Workbook ─────────────────────────────────────────────────────────────
        wb = Workbook()

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 1 — Payment Schedule
        # ═══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Payment Schedule'
        ws1.sheet_view.showGridLines = False

        # ── Cover block (rows 1-12) ───────────────────────────────────────────
        TOTAL_COLS = 30   # widest data section width in columns

        def _cover_row(ws, row, label, value, label_col=1, val_col=3,
                       label_font=None, val_font=None):
            lc = ws.cell(row=row, column=label_col, value=label)
            lc.font      = label_font or _font(bold=True, color='0D1F3C', size=10)
            lc.alignment = _align('left')
            vc = ws.cell(row=row, column=val_col, value=value)
            vc.font      = val_font or _font(bold=False, color='1A1A2E', size=10)
            vc.alignment = _align('left')

        # Title banner — merged across all columns
        ws1.merge_cells(start_row=1, start_column=1,
                         end_row=2,   end_column=TOTAL_COLS)
        title_cell = ws1.cell(row=1, column=1,
                               value='PAYROLL BULK PAYMENT FILE')
        title_cell.fill      = _fill(NAVY)
        title_cell.font      = Font(name='Calibri', bold=True, color=WHITE,
                                    size=18)
        title_cell.alignment = _align('center', 'center')
        ws1.row_dimensions[1].height = 36
        ws1.row_dimensions[2].height = 6   # spacer within merge

        # Sub-header: company name + document class
        ws1.merge_cells(start_row=3, start_column=1,
                         end_row=3,   end_column=TOTAL_COLS)
        sub = ws1.cell(row=3, column=1, value=org.name.upper())
        sub.fill      = _fill(GOLD)
        sub.font      = Font(name='Calibri', bold=True, color=NAVY, size=12)
        sub.alignment = _align('center', 'center')
        ws1.row_dimensions[3].height = 22

        # Blank gap row
        ws1.row_dimensions[4].height = 6

        # Meta fields (two-column pairs side by side)
        meta_pairs = [
            ('Pay Period',              period_label,
             'Run Reference',           run.run_number),
            ('Run Status',              run.status.upper(),
             'Payment Date',            payment_date),
            ('Currency',                curr,
             'Total Employees (Ready)', len(ready)),
            ('Total Employees (Exceptions)', len(exceptions),
             'Total Net Transfer',      tot_net),
            ('Processed By',            processed_by,
             'Approved By',             approved_by),
            ('Company Address',         org.address or '—',
             'Tax ID (TIN/VAT)',         org.tax_id or '—'),
            ('Generated On',            generated_on,
             'CONFIDENTIAL',
             'For authorised recipients only'),
        ]
        for i, (l1, v1, l2, v2) in enumerate(meta_pairs, start=5):
            row = i
            ws1.row_dimensions[row].height = 17
            # Left pair
            lc = ws1.cell(row=row, column=1, value=l1)
            lc.font = _font(bold=True, color=NAVY, size=9)
            lc.fill = _fill(MID)
            lc.alignment = _align('left')
            lc.border = _border()
            vc = ws1.cell(row=row, column=3, value=v1)
            vc.font = _font(bold=False, color='1A1A2E', size=9)
            vc.alignment = _align('left')
            vc.border = _border()
            ws1.merge_cells(start_row=row, start_column=3,
                             end_row=row,   end_column=9)
            if l1 == 'Total Net Transfer':
                vc.number_format = MONEY_FMT
                vc.value = tot_net
            # Right pair
            lc2 = ws1.cell(row=row, column=11, value=l2)
            lc2.font = _font(bold=True, color=NAVY, size=9)
            lc2.fill = _fill(MID)
            lc2.alignment = _align('left')
            lc2.border = _border()
            vc2 = ws1.cell(row=row, column=13, value=v2)
            vc2.font = _font(bold=False, color='1A1A2E', size=9)
            vc2.alignment = _align('left')
            vc2.border = _border()
            ws1.merge_cells(start_row=row, start_column=13,
                             end_row=row,   end_column=20)

        # Divider row before table
        div_row = len(meta_pairs) + 5 + 1
        ws1.row_dimensions[div_row].height = 6

        # ── Column headers (row after divider) ───────────────────────────────
        hdr_row = div_row + 1
        ws1.row_dimensions[hdr_row].height = 42

        HEADERS = [
            # identity
            'S/N', 'Employee ID', 'Full Name', 'Job Title', 'Department',
            # bank
            'Bank Name', 'Account Number\n(NUBAN)', 'Account Name', 'Bank\nCode (CBN)',
            # earnings
            'Basic\nSalary', 'Housing\nAllowance', 'Transport\nAllowance',
            'Leave\nAllowance', 'Other\nAllowances', 'Gross\nSalary',
            'Bonus', 'Overtime', 'Total\nEarnings',
            # deductions
            'Employee\nPension (8%)', 'NHF\n(2.5%)', 'NSITF\n(1%)', 'PAYE\nTax',
            'Attendance\nDeduction', 'Loan\nRepayment', 'Penalty\nDeduction',
            'Total\nDeductions',
            # net
            f'NET PAY\n({curr})',
            # statutory IDs
            "Employee's\nPFA", 'RSA PIN\n(PFA No.)', 'TIN',
            # narration
            'Bank Narration',
        ]
        MONEY_COLS = set(range(10, 28))  # 1-indexed columns that hold currency

        for col, hdr in enumerate(HEADERS, start=1):
            c = ws1.cell(row=hdr_row, column=col, value=hdr)
            c.fill      = _fill(NAVY)
            c.font      = _font(bold=True, color=WHITE, size=9)
            c.alignment = _align('center', 'center', wrap=True)
            c.border    = _border()

        # ── Data rows ────────────────────────────────────────────────────────
        first_data = hdr_row + 1
        for sn, p in enumerate(ready, start=1):
            emp   = p.employee
            drow  = first_data + sn - 1
            ws1.row_dimensions[drow].height = 16
            bg = LIGHT if sn % 2 == 0 else WHITE
            full_name   = f"{emp.first_name} {emp.last_name}".strip()
            narration   = (f"SALARY/{MONTHS[run.period_month][:3].upper()}"
                           f"-{run.period_year}/{emp.employee_id}")
            gross_total = _d(p.gross_salary) + _d(p.bonus_amount) + _d(p.overtime_amount)

            row_data = [
                sn,
                emp.employee_id,
                full_name,
                emp.job_title,
                emp.department or '',
                emp.bank_name or '',
                emp.account_number or '',
                emp.account_name or full_name,
                emp.bank_code or '',
                _d(p.basic_salary),
                _d(p.housing_allowance),
                _d(p.transport_allowance),
                _d(p.leave_allowance),
                _d(p.other_allowances),
                _d(p.gross_salary),
                _d(p.bonus_amount),
                _d(p.overtime_amount),
                gross_total,
                _d(p.employee_pension),
                _d(p.nhf),
                _d(p.nsitf),
                _d(p.paye_tax),
                _d(p.attendance_deduction),
                _d(p.loan_deductions),
                _d(p.penalty_deductions),
                _d(p.total_deductions),
                _d(p.net_salary),
                emp.pfa_name or '',
                emp.pfa_number or '',
                emp.tin or '',
                narration,
            ]
            for col, val in enumerate(row_data, start=1):
                c = ws1.cell(row=drow, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.font      = Font(name='Calibri', size=9, color='1A1A2E')
                if col in MONEY_COLS:
                    c.number_format = MONEY_FMT
                    c.alignment     = _align('right')
                elif col == 1:
                    c.alignment = _align('center')
                else:
                    c.alignment = _align('left')

        # ── Totals row ───────────────────────────────────────────────────────
        tot_row = first_data + len(ready)
        ws1.row_dimensions[tot_row].height = 18
        totals_map = {
            1:  'TOTALS',
            10: tot_basic,   11: tot_housing,  12: tot_transport,
            13: tot_leave,   14: tot_other,    15: tot_gross,
            16: tot_bonus,   17: tot_overtime, 18: tot_earnings,
            19: tot_pen_emp, 20: tot_nhf,      21: tot_nsitf,
            22: tot_paye,    23: tot_att,      24: tot_loan,
            25: tot_penalty, 26: tot_deduct,   27: tot_net,
        }
        for col in range(1, len(HEADERS) + 1):
            c = ws1.cell(row=tot_row, column=col,
                          value=totals_map.get(col, ''))
            c.fill      = _fill(GOLD)
            c.font      = Font(name='Calibri', bold=True, color=NAVY, size=9)
            c.border    = _border('medium')
            if col in MONEY_COLS:
                c.number_format = MONEY_FMT
                c.alignment     = _align('right')
            elif col == 1:
                c.alignment = _align('center')
            else:
                c.alignment = _align('left')

        # Confidentiality footer
        footer_row = tot_row + 2
        ws1.merge_cells(start_row=footer_row, start_column=1,
                         end_row=footer_row,   end_column=TOTAL_COLS)
        fc = ws1.cell(row=footer_row, column=1,
                       value=(
                           'CONFIDENTIAL — This document contains sensitive payroll '
                           'and banking information. Authorised recipients only. '
                           f'Generated by {org.name} on {generated_on}.'
                       ))
        fc.font      = _font(italic=True, color='6B7280', size=8)
        fc.alignment = _align('center')

        # Column widths (Sheet 1)
        _set_col_widths(ws1, [
            5,   # S/N
            12,  # Employee ID
            24,  # Full Name
            20,  # Job Title
            18,  # Department
            22,  # Bank Name
            18,  # Account Number
            24,  # Account Name
            10,  # Bank Code
            14,  # Basic
            14,  # Housing
            14,  # Transport
            14,  # Leave
            14,  # Other
            14,  # Gross
            12,  # Bonus
            12,  # Overtime
            14,  # Total Earnings
            14,  # Pension emp
            12,  # NHF
            12,  # NSITF
            13,  # PAYE
            14,  # Attendance
            13,  # Loan
            13,  # Penalty
            15,  # Total Deductions
            16,  # NET PAY
            22,  # PFA
            18,  # RSA PIN
            16,  # TIN
            38,  # Narration
        ])
        _freeze(ws1, f'A{hdr_row + 1}')

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 2 — Statutory Remittance Summary
        # ═══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Statutory Summary')
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells('A1:E1')
        t2 = ws2.cell(row=1, column=1, value='STATUTORY REMITTANCE SUMMARY')
        t2.fill      = _fill(NAVY)
        t2.font      = Font(name='Calibri', bold=True, color=WHITE, size=14)
        t2.alignment = _align('center', 'center')
        ws2.row_dimensions[1].height = 32

        ws2.merge_cells('A2:E2')
        s2 = ws2.cell(row=2, column=1,
                       value=f"{org.name}  ·  {period_label}  ·  Ref: {run.run_number}")
        s2.fill      = _fill(GOLD)
        s2.font      = Font(name='Calibri', bold=True, color=NAVY, size=10)
        s2.alignment = _align('center', 'center')
        ws2.row_dimensions[2].height = 18

        # Table header
        for col, hdr in enumerate(
            ['Statutory Obligation', 'Rate / Basis', f'Amount ({curr})',
             'Remittance Deadline', 'Remit To'],
            start=1,
        ):
            c = ws2.cell(row=4, column=col, value=hdr)
            c.fill      = _fill(NAVY)
            c.font      = _font(bold=True, color=WHITE, size=10)
            c.alignment = _align('center', 'center', wrap=True)
            c.border    = _border()
        ws2.row_dimensions[4].height = 28

        tot_itf = _d(run.total_itf)
        tot_benefit_emp = sum(_d(p.benefit_deductions) for p in ready)
        tot_benefit_er = _d(run.total_benefits_employer)

        stat_data = [
            # Employee deductions (withheld from pay — employer still remits them)
            ('Employee Pension Contribution',
             '8% of (Basic + Housing + Transport)',
             tot_pen_emp,
             '7th of following month',
             'Pension Fund Administrator (PFA)'),
            ('National Housing Fund (NHF)',
             '2.5% of Basic Salary',
             tot_nhf,
             '1st week of following month',
             'Federal Mortgage Bank of Nigeria (FMBN)'),
            ('PAYE Tax (Employee)',
             'NTA 2025 progressive bands',
             tot_paye,
             '10th of following month',
             'State Internal Revenue Service of each employee\'s residence'),
            ('Employee Benefit Premiums',
             'Employee share of HMO / group life',
             tot_benefit_emp,
             '1st of following month',
             'Benefit providers'),
            # Employer obligations (additional cost to the company)
            ('Employer Pension Contribution',
             '10% of (Basic + Housing + Transport)',
             tot_pen_er,
             '7th of following month',
             'Pension Fund Administrator (PFA)'),
            ('NSITF Contribution (Employer)',
             '1% of Gross Salary — employer-borne',
             tot_nsitf,
             '1st week of following month',
             'NSITF Board'),
            ('ITF Training Levy (Employer)',
             '1% of annual payroll — employer-borne',
             tot_itf,
             '1 April of following year',
             'Industrial Training Fund (ITF)'),
            ('Employer Benefit Premiums',
             'Employer share of HMO / group life',
             tot_benefit_er,
             '1st of following month',
             'Benefit providers'),
            ('Net Salary Bank Transfer',
             'Total net pay for all employees',
             tot_net,
             payment_date,
             'Employees\' bank accounts via NIBSS NIP'),
        ]
        stat_data = [row for row in stat_data if row[2]]

        for i, (obligation, rate, amount, deadline, remit_to) in \
                enumerate(stat_data, start=5):
            bg = LIGHT if i % 2 == 0 else WHITE
            ws2.row_dimensions[i].height = 18
            is_employer = 'Employer' in obligation or 'Transfer' in obligation
            row_vals = [obligation, rate, amount, deadline, remit_to]
            for col, val in enumerate(row_vals, start=1):
                c = ws2.cell(row=i, column=col, value=val)
                c.fill      = _fill(bg)
                c.font      = Font(
                    name='Calibri', size=9, color='1A1A2E',
                    bold=(col == 3 and is_employer),
                )
                c.border    = _border()
                c.alignment = _align('left' if col != 3 else 'right',
                                     wrap=(col == 1))
                if col == 3:
                    c.number_format = MONEY_FMT

        # Grand total — every naira the employer must fund for this run.
        # The employee pension contribution was previously omitted here, which
        # understated the funding requirement by 8% of emoluments.
        grand_row = len(stat_data) + 5
        ws2.row_dimensions[grand_row].height = 20
        grand_total = (
            tot_net + tot_paye + tot_pen_emp + tot_pen_er + tot_nhf
            + tot_nsitf + tot_itf + tot_benefit_emp + tot_benefit_er
        )
        labels = ['TOTAL EMPLOYER FUNDING REQUIREMENT',
                  '(Net Pay + PAYE + Pension employee & employer + NHF + NSITF + ITF + Benefits)',
                  grand_total, '', '']
        for col, val in enumerate(labels, start=1):
            c = ws2.cell(row=grand_row, column=col, value=val)
            c.fill      = _fill(GOLD)
            c.font      = Font(name='Calibri', bold=True, color=NAVY, size=10)
            c.border    = _border('medium')
            c.alignment = _align('right' if col == 3 else 'left')
            if col == 3:
                c.number_format = MONEY_FMT

        _set_col_widths(ws2, [38, 38, 18, 28, 42])

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 3 — Exceptions
        # ═══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Exceptions')
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells('A1:L1')
        t3 = ws3.cell(row=1, column=1,
                       value='EXCEPTIONS — INCOMPLETE BANK DETAILS')
        t3.fill      = PatternFill('solid', fgColor='8B0000')
        t3.font      = Font(name='Calibri', bold=True, color=WHITE, size=14)
        t3.alignment = _align('center', 'center')
        ws3.row_dimensions[1].height = 30

        ws3.merge_cells('A2:L2')
        s3 = ws3.cell(row=2, column=1,
                       value=(f"{len(exceptions)} employee(s) excluded from bank transfer — "
                              "update their bank details and re-run."))
        s3.fill      = PatternFill('solid', fgColor='FFE4E4')
        s3.font      = Font(name='Calibri', italic=True, color=RED_TX, size=10)
        s3.alignment = _align('center')
        ws3.row_dimensions[2].height = 18

        exc_headers = [
            'S/N', 'Employee ID', 'Full Name', 'Job Title', 'Department',
            'Bank Name', 'Account Number', 'Account Name', 'Bank Code',
            f'Net Pay Due ({curr})', 'Missing Fields', 'Action Required',
        ]
        for col, hdr in enumerate(exc_headers, start=1):
            c = ws3.cell(row=4, column=col, value=hdr)
            c.fill      = PatternFill('solid', fgColor='8B0000')
            c.font      = _font(bold=True, color=WHITE, size=9)
            c.alignment = _align('center', 'center', wrap=True)
            c.border    = _border()
        ws3.row_dimensions[4].height = 28

        if exceptions:
            for sn, p in enumerate(exceptions, start=1):
                emp  = p.employee
                drow = sn + 4
                ws3.row_dimensions[drow].height = 16
                missing = []
                if not (emp.account_number or '').strip():
                    missing.append('Account Number')
                if not (emp.bank_code or '').strip():
                    missing.append('Bank Code')
                if not (emp.account_name or '').strip():
                    missing.append('Account Name')
                exc_row = [
                    sn,
                    emp.employee_id,
                    f"{emp.first_name} {emp.last_name}".strip(),
                    emp.job_title,
                    emp.department or '',
                    emp.bank_name or 'NOT SET',
                    emp.account_number or 'NOT SET',
                    emp.account_name or 'NOT SET',
                    emp.bank_code or 'NOT SET',
                    _d(p.net_salary),
                    ', '.join(missing),
                    'Update bank details in Payroll → Employees',
                ]
                for col, val in enumerate(exc_row, start=1):
                    c = ws3.cell(row=drow, column=col, value=val)
                    c.fill      = _fill(RED_BG)
                    c.font      = Font(name='Calibri', size=9, color=RED_TX)
                    c.border    = _border()
                    c.alignment = _align(
                        'right' if col == 10 else 'center' if col == 1 else 'left'
                    )
                    if col == 10:
                        c.number_format = MONEY_FMT
        else:
            ws3.merge_cells('A5:L5')
            ok = ws3.cell(row=5, column=1,
                           value='✓  No exceptions — all employees have complete bank details.')
            ok.fill      = PatternFill('solid', fgColor='E6F9EC')
            ok.font      = Font(name='Calibri', color='166534', bold=True, size=10)
            ok.alignment = _align('center')

        _set_col_widths(ws3, [5, 12, 24, 20, 18, 22, 18, 24, 10, 16, 30, 36])

        # ── Serialize and respond ────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = (f'{run.run_number}-payroll-bank-file-'
                    f'{run.period_year}{run.period_month:02d}.xlsx')
        response = HttpResponse(
            buf.read(),
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def register(self, request):
        """
        GET /payroll/runs/register/?year=<int>

        Payroll register with month-on-month variance. Pure aggregation over
        PayrollRun's existing denormalized total columns — no per-payslip
        iteration.
        """
        from decimal import Decimal

        org = self._get_organisation()
        year = int(request.query_params.get('year') or PayrollRun.objects.filter(
            organisation=org).order_by('-period_year').values_list('period_year', flat=True).first() or 0)

        runs = list(
            PayrollRun.objects.filter(organisation=org, period_year=year, run_type=PayrollRun.REGULAR)
            .order_by('period_month')
            .values(
                'id', 'run_number', 'period_month', 'status',
                'total_gross', 'total_net', 'total_paye', 'total_deductions',
            )
        )
        rows = []
        prev_net = None
        for r in runs:
            net = Decimal(str(r['total_net'] or 0))
            variance_pct = None
            if prev_net is not None and prev_net != 0:
                variance_pct = round(float((net - prev_net) / prev_net) * 100, 2)
            rows.append({
                **r, 'id': str(r['id']),
                'variance_from_prior_month_pct': variance_pct,
            })
            prev_net = net
        return Response({'year': year, 'rows': rows})

    @action(detail=False, methods=['get'])
    def annual_paye_reconciliation(self, request):
        """
        GET /payroll/runs/annual_paye_reconciliation/?year=<int>

        Read-only: compares actual monthly PayslipLine.paye_tax sums per
        employee for the tax year against the correct annual tax on their
        actual annual taxable income, surfacing the variance. Corrections
        apply via the existing PayrollAdjustment CORRECTION type — this
        endpoint only reports, it never writes.
        """
        from datetime import date as _date
        from decimal import Decimal

        from django.db.models import Sum

        CENTS = Decimal('0.01')
        org = self._get_organisation()
        year = int(request.query_params.get('year') or _date.today().year)

        rows = (
            PayslipLine.objects
            .filter(organisation=org, payroll_run__period_year=year)
            .values('employee_id', 'employee__first_name', 'employee__last_name', 'employee__employee_id')
            .annotate(
                actual_taxable=Sum('taxable_income'),
                actual_paye_withheld=Sum('paye_tax'),
            )
            .order_by('employee__last_name')
        )
        results = []
        for r in rows:
            actual_taxable = Decimal(str(r['actual_taxable'] or 0))
            correct_annual_paye = PayrollService.calculate_annual_paye(actual_taxable).quantize(CENTS, rounding='ROUND_HALF_UP')
            withheld = Decimal(str(r['actual_paye_withheld'] or 0))
            variance = withheld - correct_annual_paye
            results.append({
                'employee_id': str(r['employee_id']),
                'employee_code': r['employee__employee_id'],
                'employee_name': f"{r['employee__first_name']} {r['employee__last_name']}".strip(),
                'actual_taxable_income': actual_taxable,
                'actual_paye_withheld': withheld,
                'correct_annual_paye': correct_annual_paye,
                'variance': variance,
                'variance_direction': (
                    'over-withheld' if variance > 0 else 'under-withheld' if variance < 0 else 'exact'
                ),
            })
        return Response({'year': year, 'rows': results})

    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """GET /payroll/runs/pending_approvals/ — runs awaiting approval (for notification badge)."""
        org = self._get_organisation()
        runs = PayrollRun.objects.filter(
            organisation=org,
            status=PayrollRun.PROCESSING,
            submitted_for_approval=True,
        ).values('id', 'run_number', 'period_year', 'period_month', 'submitted_by__email')
        return Response(list(runs))

    @action(detail=True, methods=['post'])
    def send_payslips(self, request, pk=None):
        """
        POST /payroll/runs/{id}/send_payslips/

        Body: {"payslips": [{"id": "...", "pdf_base64": "..."}], "subject": "..."}

        Mirrors the invoice-email path: the client renders each PDF and posts it
        back, the server attaches and sends. Every attempt is written to
        PayslipDelivery — issuing a payslip is a compliance act and needs a trail.
        """
        import base64
        from email import encoders
        from email.mime.base import MIMEBase

        from django.core.mail import EmailMultiAlternatives

        run = self.get_object()
        items = request.data.get('payslips') or []
        if not isinstance(items, list) or not items:
            return Response({'error': 'Provide a non-empty payslips array.'}, status=400)

        by_id = {
            str(p.id): p
            for p in run.payslips.select_related('employee').all()
        }
        org_name = run.organisation.name
        period = f"{run.period_year}-{run.period_month:02d}"
        sent, failed, skipped = 0, 0, 0
        results = []

        for item in items:
            payslip = by_id.get(str(item.get('id')))
            if payslip is None:
                continue
            employee = payslip.employee
            recipient = (employee.email or '').strip()
            if not recipient:
                PayslipDelivery.objects.create(
                    organisation=run.organisation, payslip=payslip,
                    channel=PayslipDelivery.EMAIL, recipient='',
                    status=PayslipDelivery.SKIPPED,
                    error='No email address on file', sent_by=request.user,
                )
                skipped += 1
                results.append({'id': str(payslip.id), 'status': 'skipped',
                                'reason': 'No email address on file'})
                continue

            try:
                subject = request.data.get('subject') or f"Payslip for {period} — {org_name}"
                body = (
                    f"Dear {employee.first_name},\n\n"
                    f"Your payslip for {period} is attached.\n\n"
                    f"Net pay: {payslip.net_salary}\n\n"
                    f"{org_name}"
                )
                msg = EmailMultiAlternatives(subject, body, to=[recipient])
                pdf_b64 = item.get('pdf_base64')
                if pdf_b64:
                    if ',' in pdf_b64[:64]:
                        pdf_b64 = pdf_b64.split(',', 1)[1]
                    part = MIMEBase('application', 'pdf')
                    part.set_payload(base64.b64decode(pdf_b64))
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition', 'attachment',
                        filename=f"Payslip-{employee.employee_id}-{period}.pdf",
                    )
                    msg.attach(part)
                msg.send(fail_silently=False)

                PayslipDelivery.objects.create(
                    organisation=run.organisation, payslip=payslip,
                    channel=PayslipDelivery.EMAIL, recipient=recipient,
                    status=PayslipDelivery.SENT, sent_by=request.user,
                )
                sent += 1
                results.append({'id': str(payslip.id), 'status': 'sent'})
            except Exception as exc:
                PayslipDelivery.objects.create(
                    organisation=run.organisation, payslip=payslip,
                    channel=PayslipDelivery.EMAIL, recipient=recipient,
                    status=PayslipDelivery.FAILED, error=str(exc)[:500],
                    sent_by=request.user,
                )
                failed += 1
                results.append({'id': str(payslip.id), 'status': 'failed', 'reason': str(exc)[:200]})

        return Response({
            'sent': sent, 'failed': failed, 'skipped': skipped, 'results': results,
        })

    @action(detail=True, methods=['post'])
    def send_payslips_server_rendered(self, request, pk=None):
        """
        POST /payroll/runs/{id}/send_payslips_server_rendered/
        Body (optional): { employee_ids: [...] }

        Queues send_payslips_async — server renders and emails every payslip
        PDF itself, no client rendering required. Falls back to a synchronous
        call if Celery is not configured/eager in this environment.
        """
        run = self.get_object()
        employee_ids = request.data.get('employee_ids') or None
        from .tasks import send_payslips_async

        try:
            async_result = send_payslips_async.delay(str(run.id), employee_ids)
            return Response({'queued': True, 'task_id': async_result.id})
        except Exception:
            # No broker available (e.g. local/dev without Celery running) —
            # run synchronously so the feature still works end-to-end.
            result = send_payslips_async(str(run.id), employee_ids)
            return Response({'queued': False, **result})

    @action(detail=True, methods=['get'], url_path='payslip-pdf/(?P<employee_id>[^/.]+)')
    def payslip_pdf(self, request, pk=None, employee_id=None):
        """GET /payroll/runs/{id}/payslip-pdf/{employee_id}/ — one payslip as a downloadable PDF."""
        from .pdf import build_payslip_pdf

        run = self.get_object()
        try:
            payslip = run.payslips.select_related('employee', 'organisation').get(employee_id=employee_id)
        except PayslipLine.DoesNotExist:
            return Response({'error': 'Payslip not found for this employee in this run.'}, status=404)
        pdf_bytes = build_payslip_pdf(payslip)
        filename = f"Payslip-{payslip.employee.employee_id}-{run.period_year}{run.period_month:02d}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'])
    def deliveries(self, request, pk=None):
        """GET /payroll/runs/{id}/deliveries/ — payslip delivery audit trail."""
        run = self.get_object()
        qs = PayslipDelivery.objects.filter(
            payslip__payroll_run=run
        ).select_related('payslip__employee')
        return Response(PayslipDeliverySerializer(qs, many=True).data)


class EmployeeTaxProfileViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """GET/PUT /payroll/tax-profiles/ — per-employee tax relief overrides."""
    serializer_class = EmployeeTaxProfileSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        org = self._get_organisation()
        return EmployeeTaxProfile.objects.filter(organisation=org).select_related('employee')

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=False, methods=['get', 'put', 'patch'], url_path='by_employee/(?P<employee_id>[^/.]+)')
    def by_employee(self, request, employee_id=None):
        """GET/PUT /payroll/tax-profiles/by_employee/{employee_id}/ — upsert profile for one employee."""
        org = self._get_organisation()
        profile, _ = EmployeeTaxProfile.objects.get_or_create(
            organisation=org, employee_id=employee_id,
            defaults={'nhf_enrolled': True, 'voluntary_pension': 0, 'life_assurance_premium': 0, 'paye_exempt': False},
        )
        if request.method == 'GET':
            return Response(EmployeeTaxProfileSerializer(profile).data)
        serializer = EmployeeTaxProfileSerializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class StatutoryRemittanceViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """
    GET/PATCH /payroll/remittances/ — every statutory and benefit obligation.

    Rows are generated by the payroll run, so there is no create endpoint.
    """
    serializer_class = StatutoryRemittanceSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    http_method_names = ['get', 'patch', 'post', 'head', 'options']

    def get_queryset(self):
        org = self._get_organisation()
        qs = (
            StatutoryRemittance.objects
            .filter(organisation=org)
            .select_related('payroll_run', 'tax_authority')
        )
        params = self.request.query_params
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('remittance_type'):
            qs = qs.filter(remittance_type=params['remittance_type'])
        if params.get('period_year'):
            qs = qs.filter(period_year=params['period_year'])
        if params.get('period_month'):
            qs = qs.filter(period_month=params['period_month'])
        if params.get('overdue') == 'true':
            from django.utils import timezone
            qs = qs.exclude(status=StatutoryRemittance.REMITTED).filter(
                due_date__lt=timezone.localdate()
            )
        return qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """GET /payroll/remittances/summary/ — headline figures for the cockpit."""
        from decimal import Decimal as _Dec
        from django.db.models import Sum
        from django.utils import timezone

        org = self._get_organisation()
        qs = StatutoryRemittance.objects.filter(organisation=org)
        today = timezone.localdate()

        outstanding_qs = qs.exclude(status=StatutoryRemittance.REMITTED)
        outstanding = sum(
            (_Dec(str(r.balance_due)) for r in outstanding_qs), _Dec('0')
        )
        overdue_qs = outstanding_qs.filter(due_date__lt=today)
        overdue = sum((_Dec(str(r.balance_due)) for r in overdue_qs), _Dec('0'))
        remitted_ytd = qs.filter(
            status=StatutoryRemittance.REMITTED, period_year=today.year,
        ).aggregate(total=Sum('amount_paid'))['total'] or _Dec('0')
        next_due = outstanding_qs.order_by('due_date').first()

        return Response({
            'outstanding': str(outstanding),
            'outstanding_count': outstanding_qs.count(),
            'overdue': str(overdue),
            'overdue_count': overdue_qs.count(),
            'remitted_ytd': str(remitted_ytd),
            'next_due_date': next_due.due_date if next_due else None,
            'next_due_recipient': (
                next_due.recipient_name or (next_due.tax_authority.name if next_due and next_due.tax_authority else '')
            ) if next_due else None,
        })

    @action(detail=True, methods=['post'])
    def mark_remitted(self, request, pk=None):
        """
        POST /payroll/remittances/{id}/mark_remitted/

        Records the payment and posts the clearing journal, so the liability
        actually leaves the balance sheet instead of accruing forever.
        """
        remittance = self.get_object()
        if remittance.status == StatutoryRemittance.REMITTED:
            return Response({'error': 'This obligation has already been remitted.'}, status=400)
        amount = request.data.get('amount_paid')
        RemittanceService.mark_remitted(
            remittance,
            amount=amount,
            reference=request.data.get('reference', ''),
            user=request.user,
        )
        remittance.refresh_from_db()
        return Response(StatutoryRemittanceSerializer(remittance).data)

    @action(detail=False, methods=['get'])
    def schedule(self, request):
        """
        GET /payroll/remittances/schedule/?type=pension&year=&month=

        Per-recipient filing schedule. A PFA will not accept a blended file, so
        pension exports one sheet per PFA and PAYE one per State IRS.
        """
        org = self._get_organisation()
        r_type = request.query_params.get('type', StatutoryRemittance.PENSION)
        year = request.query_params.get('year')
        month = request.query_params.get('month')

        runs = PayrollRun.objects.filter(organisation=org)
        if year:
            runs = runs.filter(period_year=year)
        if month:
            runs = runs.filter(period_month=month)
        payslips = (
            PayslipLine.objects
            .filter(organisation=org, payroll_run__in=runs)
            .select_related('employee', 'tax_authority', 'payroll_run')
        )

        groups: dict = {}
        for slip in payslips:
            emp = slip.employee
            if r_type == StatutoryRemittance.PAYE:
                key = slip.tax_authority.name if slip.tax_authority_id else 'Unassigned — set state of residence'
                amount = float(slip.paye_tax or 0)
                detail = {'tin': emp.tin, 'state': emp.get_state_of_residence_display() or ''}
            elif r_type == StatutoryRemittance.PENSION:
                key = (emp.pfa_name or '').strip() or 'Unassigned PFA'
                amount = float(slip.employee_pension or 0) + float(slip.employer_pension or 0)
                detail = {
                    'pfa_number': emp.pfa_number, 'pension_pin': emp.pension_pin,
                    'employee_share': float(slip.employee_pension or 0),
                    'employer_share': float(slip.employer_pension or 0),
                }
            elif r_type == StatutoryRemittance.NHF:
                key = 'Federal Mortgage Bank of Nigeria (FMBN)'
                amount = float(slip.nhf or 0)
                detail = {}
            else:
                key = 'All employees'
                amount = float(slip.nsitf or 0)
                detail = {}
            if amount <= 0:
                continue
            groups.setdefault(key, {'recipient': key, 'total': 0.0, 'employees': []})
            groups[key]['total'] += amount
            groups[key]['employees'].append({
                'employee_id': emp.employee_id,
                'name': f"{emp.first_name} {emp.last_name}",
                'gross': float(slip.gross_salary or 0),
                'amount': round(amount, 2),
                **detail,
            })

        for group in groups.values():
            group['total'] = round(group['total'], 2)
            group['count'] = len(group['employees'])
        return Response({
            'type': r_type,
            'groups': sorted(groups.values(), key=lambda g: -g['total']),
        })


class TaxAuthorityViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """GET/PATCH /payroll/tax-authorities/ — the State IRS registry."""
    serializer_class = TaxAuthoritySerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        if not TaxAuthority.objects.filter(organisation=org).exists():
            TaxAuthorityService.seed(org)
        return TaxAuthority.objects.filter(organisation=org)

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


class CompensationRecordViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """
    GET/POST /payroll/compensation/ — effective-dated salary history.

    Writing a record here is what makes a backdated raise auditable and lets the
    engine compute arrears.
    """
    serializer_class = CompensationRecordSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        org = self._get_organisation()
        qs = CompensationRecord.objects.filter(organisation=org).select_related('employee')
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs

    def perform_create(self, serializer):
        org = self._get_organisation()
        employee = serializer.validated_data['employee']
        if employee.organisation_id != org.id:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('That employee belongs to another organisation.')
        record = CompensationService.record_change(
            employee=employee,
            effective_date=serializer.validated_data['effective_date'],
            reason=serializer.validated_data.get('reason', CompensationRecord.ADJUSTMENT),
            notes=serializer.validated_data.get('notes', ''),
            **{
                f: serializer.validated_data.get(f)
                for f in CompensationService.COMPONENTS
                if serializer.validated_data.get(f) is not None
            },
        )
        serializer.instance = record


class PayrollAdjustmentViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """GET/POST /payroll/adjustments/ — arrears and back-pay."""
    serializer_class = PayrollAdjustmentSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        org = self._get_organisation()
        qs = PayrollAdjustment.objects.filter(organisation=org).select_related('employee')
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        if self.request.query_params.get('status'):
            qs = qs.filter(status=self.request.query_params['status'])
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        adjustment = self.get_object()
        if adjustment.status != PayrollAdjustment.PENDING:
            return Response({'error': 'Only pending adjustments can be cancelled.'}, status=400)
        adjustment.status = PayrollAdjustment.CANCELLED
        adjustment.save(update_fields=['status'])
        return Response(PayrollAdjustmentSerializer(adjustment).data)

    @action(detail=False, methods=['post'])
    def request_encashment(self, request):
        """
        POST /payroll/adjustments/request_encashment/
        Body: { employee: <id>, leave_type: <id>, days: <decimal>, reason: '' }

        Creates a pending PayrollAdjustment(ENCASHMENT), mirroring the existing
        Requests-tab approval pattern — it stays PENDING until picked up by a
        payroll run, same as arrears/back-pay.
        """
        org = self._get_organisation()
        try:
            employee = Employee.objects.get(organisation=org, id=request.data.get('employee'))
        except (Employee.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Employee not found in this organisation.'}, status=404)
        try:
            leave_type = LeaveType.objects.get(organisation=org, id=request.data.get('leave_type'))
        except (LeaveType.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Leave type not found in this organisation.'}, status=404)
        try:
            adjustment = LeaveEncashmentService.request_encashment(
                employee, leave_type, request.data.get('days'), request.data.get('reason', ''),
            )
        except ValueError as exc:
            return Response({'error': str(exc)}, status=400)
        return Response(PayrollAdjustmentSerializer(adjustment).data, status=201)


class PayrollSettingsViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """GET/PATCH /payroll/settings/ — org-level payroll configuration."""
    serializer_class = PayrollSettingsSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    http_method_names = ['get', 'patch', 'head', 'options']

    def get_queryset(self):
        org = self._get_organisation()
        get_settings(org)
        return PayrollSettings.objects.filter(organisation=org)

    @action(detail=False, methods=['get', 'patch'])
    def current(self, request):
        org = self._get_organisation()
        settings_row = get_settings(org)
        if request.method == 'GET':
            return Response(PayrollSettingsSerializer(settings_row).data)
        serializer = PayrollSettingsSerializer(settings_row, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# Leave
# ══════════════════════════════════════════════════════════════════════════════

class LeaveTypeViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = LeaveTypeSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        if not LeaveType.objects.filter(organisation=org).exists():
            LeaveService.seed_defaults(org)
        return LeaveType.objects.filter(organisation=org)

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


class PublicHolidayViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    """CRUD /payroll/public-holidays/ — org-recognised public holidays."""
    serializer_class = PublicHolidaySerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = PublicHoliday.objects.filter(organisation=org)
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(date__year=year)
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())

    @action(detail=False, methods=['post'])
    def seed(self, request):
        """POST /payroll/public-holidays/seed/ — seed fixed-date holidays for a year."""
        org = self._get_organisation()
        try:
            year = int(request.data.get('year'))
        except (TypeError, ValueError):
            return Response({'error': 'year is required.'}, status=400)
        qs = PublicHolidayService.seed_fixed_dates(org, year)
        return Response(PublicHolidaySerializer(qs, many=True).data)


class LeaveCarryForwardViewSet(TenantFilterMixin, viewsets.ViewSet):
    """
    Read-only preview + apply actions for year-end leave carry-forward.
    Not a ModelViewSet — this is a workflow, not a CRUD resource.
    """
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin, _PlanPayroll]

    @action(detail=False, methods=['get'])
    def preview(self, request):
        """GET /payroll/leave-carry-forward/preview/?year=<prior_year>"""
        org = self._get_organisation()
        try:
            prior_year = int(request.query_params.get('year'))
        except (TypeError, ValueError):
            return Response({'error': 'year (the prior/source year) is required.'}, status=400)
        rows = LeaveService.carry_forward_preview(org, prior_year)
        return Response(rows)

    @action(detail=False, methods=['post'])
    def apply(self, request):
        """POST /payroll/leave-carry-forward/apply/ Body: { prior_year, new_year }"""
        org = self._get_organisation()
        try:
            prior_year = int(request.data.get('prior_year'))
            new_year = int(request.data.get('new_year'))
        except (TypeError, ValueError):
            return Response({'error': 'prior_year and new_year are required.'}, status=400)
        updated = LeaveService.carry_forward_year_end(org, prior_year, new_year)
        return Response({'updated': updated, 'prior_year': prior_year, 'new_year': new_year})


class LeaveBalanceViewSet(TenantFilterMixin, viewsets.ReadOnlyModelViewSet):
    serializer_class = LeaveBalanceSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = (
            LeaveBalance.objects.filter(organisation=org)
            .select_related('employee', 'leave_type')
        )
        params = self.request.query_params
        if params.get('employee'):
            qs = qs.filter(employee_id=params['employee'])
        if params.get('year'):
            qs = qs.filter(year=params['year'])
        return qs

    @action(detail=False, methods=['post'])
    def accrue(self, request):
        """POST /payroll/leave-balances/accrue/ — run monthly accrual on demand."""
        from datetime import date as _date
        org = self._get_organisation()
        today = _date.today()
        year = int(request.data.get('year') or today.year)
        month = int(request.data.get('month') or today.month)
        updated = LeaveService.accrue_month(org, year, month)
        return Response({'updated': updated, 'year': year, 'month': month})


class LeaveRequestViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = LeaveRequestSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = (
            LeaveRequest.objects.filter(organisation=org)
            .select_related('employee', 'leave_type', 'decided_by')
        )
        params = self.request.query_params
        if params.get('employee'):
            qs = qs.filter(employee_id=params['employee'])
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('year'):
            qs = qs.filter(start_date__year=params['year'])
        return qs

    def create(self, request, *args, **kwargs):
        """
        HR-facing leave request creation — warn-and-allow, never hard-block.

        Unlike the ESS-facing endpoint (ess_views.MeLeaveRequestViewSet, which
        still hard-blocks), HR raising a request on an employee's behalf may
        knowingly over-book:
          - Tier 1 (soft warn): days exceed available_days but the balance
            stays >= 0 after — allowed, flagged, no reason required.
          - Tier 2 (hard warn): the request would push the balance negative —
            allowed, but a non-blank ``reason`` is mandatory (400 if blank).
        """
        from decimal import Decimal as _Dec

        org = self._get_organisation()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        employee = serializer.validated_data['employee']
        leave_type = serializer.validated_data['leave_type']
        days = LeaveRequest.working_days_between(
            serializer.validated_data['start_date'], serializer.validated_data['end_date']
        )

        balance = LeaveService.get_or_create_balance(
            employee, leave_type, serializer.validated_data['start_date'].year
        )
        is_overbooked = False
        overbooked_days = _Dec('0')
        if leave_type.is_paid and days > balance.available_days:
            is_overbooked = True
            overbooked_days = days - balance.available_days
            projected_balance = balance.available_days - days
            if projected_balance < 0:
                reason = (request.data.get('reason') or '').strip()
                if not reason:
                    return Response(
                        {'error': (
                            f'This request of {days} days would take the balance to '
                            f'{projected_balance} (negative). A reason is required to proceed.'
                        )},
                        status=400,
                    )

        # Route the approval to the employee's manager where one is recorded.
        approver = employee.manager.user if (employee.manager and employee.manager.user_id) else None
        instance = serializer.save(
            organisation=org, days=days, approver=approver,
            status=LeaveRequest.PENDING if leave_type.requires_approval else LeaveRequest.APPROVED,
            is_overbooked=is_overbooked,
            overbooked_by=request.user if is_overbooked else None,
            overbooked_days=overbooked_days,
        )
        if instance.status == LeaveRequest.PENDING:
            balance.pending_days = _Dec(str(balance.pending_days)) + days
            balance.save(update_fields=['pending_days'])
        else:
            LeaveService.approve(instance, user=self.request.user)

        headers = self.get_success_headers(serializer.data)
        return Response(
            self.get_serializer(instance).data, status=status.HTTP_201_CREATED, headers=headers,
        )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        leave_request = self.get_object()
        if leave_request.status not in [LeaveRequest.PENDING, LeaveRequest.DRAFT]:
            return Response({'error': 'Only pending requests can be approved.'}, status=400)
        LeaveService.approve(leave_request, user=request.user, note=request.data.get('note', ''))
        leave_request.refresh_from_db()
        return Response(LeaveRequestSerializer(leave_request).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        leave_request = self.get_object()
        if leave_request.status not in [LeaveRequest.PENDING, LeaveRequest.DRAFT]:
            return Response({'error': 'Only pending requests can be rejected.'}, status=400)
        LeaveService.reject(leave_request, user=request.user, note=request.data.get('note', ''))
        leave_request.refresh_from_db()
        return Response(LeaveRequestSerializer(leave_request).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        leave_request = self.get_object()
        if leave_request.status in [LeaveRequest.CANCELLED, LeaveRequest.REJECTED]:
            return Response({'error': 'This request is already closed.'}, status=400)
        LeaveService.cancel(leave_request, user=request.user)
        leave_request.refresh_from_db()
        return Response(LeaveRequestSerializer(leave_request).data)

    @action(detail=False, methods=['get'])
    def pending_count(self, request):
        org = self._get_organisation()
        count = LeaveRequest.objects.filter(
            organisation=org, status=LeaveRequest.PENDING
        ).count()
        return Response({'count': count})

    @action(detail=False, methods=['get'])
    def team_coverage(self, request):
        """
        GET /payroll/leave-requests/team_coverage/?employee=<id>&start_date=&end_date=

        Who else on this employee's team is off over the requested window —
        approved or pending requests for anyone sharing the same manager or
        department, overlapping [start_date, end_date]. One query.
        """
        org = self._get_organisation()
        employee_id = request.query_params.get('employee')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if not employee_id or not start_date or not end_date:
            return Response(
                {'error': 'employee, start_date and end_date are required.'}, status=400,
            )
        try:
            employee = Employee.objects.get(organisation=org, id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=404)

        from django.db.models import Q

        if not employee.department and not employee.manager_id:
            return Response([])

        peer_filter = Q(employee__department=employee.department) if employee.department else Q()
        if employee.manager_id:
            peer_filter |= Q(employee__manager_id=employee.manager_id)

        rows = (
            LeaveRequest.objects
            .filter(
                organisation=org,
                status__in=[LeaveRequest.APPROVED, LeaveRequest.PENDING],
                start_date__lte=end_date,
                end_date__gte=start_date,
            )
            .filter(peer_filter)
            .exclude(employee_id=employee.id)
            .select_related('employee', 'leave_type')
        )
        data = [
            {
                'employee_id': str(r.employee_id),
                'employee_name': r.employee.full_name,
                'department': r.employee.department,
                'leave_type': r.leave_type.name,
                'start_date': r.start_date,
                'end_date': r.end_date,
                'status': r.status,
            }
            for r in rows
        ]
        return Response(data)


# ══════════════════════════════════════════════════════════════════════════════
# Benefits
# ══════════════════════════════════════════════════════════════════════════════

class BenefitPlanViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = BenefitPlanSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        return BenefitPlan.objects.filter(organisation=self._get_organisation())

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


class EmployeeBenefitViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = EmployeeBenefitSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = (
            EmployeeBenefit.objects.filter(organisation=org)
            .select_related('employee', 'plan')
        )
        if self.request.query_params.get('employee'):
            qs = qs.filter(employee_id=self.request.query_params['employee'])
        return qs

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


# ══════════════════════════════════════════════════════════════════════════════
# Salary advances (earned wage access)
# ══════════════════════════════════════════════════════════════════════════════

class AdvanceRequestViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = AdvanceRequestSerializer
    permission_classes = [IsAuthenticated, IsStaff, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = AdvanceRequest.objects.filter(organisation=org).select_related('employee')
        params = self.request.query_params
        if params.get('employee'):
            qs = qs.filter(employee_id=params['employee'])
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        return qs

    def create(self, request, *args, **kwargs):
        org = self._get_organisation()
        employee_id = request.data.get('employee')
        try:
            employee = Employee.objects.get(organisation=org, id=employee_id)
        except (Employee.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Employee not found in this organisation.'}, status=404)
        try:
            advance = EWAService.request(
                employee, request.data.get('amount'), request.data.get('reason', '')
            )
        except ValueError as exc:
            return Response({'error': str(exc)}, status=400)
        return Response(AdvanceRequestSerializer(advance).data, status=201)

    @action(detail=False, methods=['get'], url_path='eligibility/(?P<employee_id>[^/.]+)')
    def eligibility(self, request, employee_id=None):
        """GET /payroll/advances/eligibility/{employee_id}/ — what may be drawn now."""
        org = self._get_organisation()
        try:
            employee = Employee.objects.get(organisation=org, id=employee_id)
        except (Employee.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Employee not found in this organisation.'}, status=404)
        info = EWAService.eligibility(employee)
        return Response({k: (str(v) if hasattr(v, 'quantize') else v) for k, v in info.items()})

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve and disburse. Gated on the organisation's own cash position —
        the underwriting signal a payroll-only platform does not hold.
        """
        from django.utils import timezone

        advance = self.get_object()
        if advance.status != AdvanceRequest.PENDING:
            return Response({'error': 'Only pending advances can be approved.'}, status=400)

        can_fund, reason = EWAService.can_employer_fund(advance.organisation, advance.amount)
        if not can_fund:
            return Response({'error': reason}, status=400)

        advance.status = AdvanceRequest.DISBURSED
        advance.decided_by = request.user
        advance.decided_at = timezone.now()
        advance.decision_note = request.data.get('note', '')
        advance.disbursed_at = timezone.now()
        advance.save(update_fields=[
            'status', 'decided_by', 'decided_at', 'decision_note', 'disbursed_at',
        ])

        try:
            from apps.accounting.services import AccountingService, safe_post_gl
            safe_post_gl(AccountingService.post_advance_journal, advance, user=request.user)
        except Exception:
            pass
        return Response(AdvanceRequestSerializer(advance).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        from django.utils import timezone

        advance = self.get_object()
        if advance.status != AdvanceRequest.PENDING:
            return Response({'error': 'Only pending advances can be rejected.'}, status=400)
        advance.status = AdvanceRequest.REJECTED
        advance.decided_by = request.user
        advance.decided_at = timezone.now()
        advance.decision_note = request.data.get('note', '')
        advance.save(update_fields=['status', 'decided_by', 'decided_at', 'decision_note'])
        return Response(AdvanceRequestSerializer(advance).data)


class AdvancePolicyViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = AdvancePolicySerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    http_method_names = ['get', 'patch', 'head', 'options']

    def get_queryset(self):
        org = self._get_organisation()
        EWAService.get_policy(org)
        return AdvancePolicy.objects.filter(organisation=org)

    @action(detail=False, methods=['get', 'patch'])
    def current(self, request):
        org = self._get_organisation()
        policy = EWAService.get_policy(org)
        if request.method == 'GET':
            return Response(AdvancePolicySerializer(policy).data)
        serializer = AdvancePolicySerializer(policy, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# Offboarding (A.3)
# ══════════════════════════════════════════════════════════════════════════════

class OffboardingCaseViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = OffboardingCaseSerializer
    permission_classes = [IsAuthenticated, IsManager, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        qs = (
            OffboardingCase.objects.filter(organisation=org)
            .select_related('employee', 'initiated_by', 'completed_by')
            .prefetch_related('checklist_items', 'exit_interview')
        )
        if self.request.query_params.get('status'):
            qs = qs.filter(status=self.request.query_params['status'])
        if self.request.query_params.get('employee'):
            qs = qs.filter(employee_id=self.request.query_params['employee'])
        return qs

    def create(self, request, *args, **kwargs):
        org = self._get_organisation()
        try:
            employee = Employee.objects.get(organisation=org, id=request.data.get('employee'))
        except (Employee.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Employee not found in this organisation.'}, status=404)
        reason = request.data.get('reason')
        valid_reasons = [c for c, _ in OffboardingCase.REASON_CHOICES]
        if reason not in valid_reasons:
            return Response({'error': f"reason must be one of: {', '.join(valid_reasons)}"}, status=400)
        last_working_day = request.data.get('last_working_day')
        if not last_working_day:
            return Response({'error': 'last_working_day is required.'}, status=400)
        case = OffboardingService.create_case(
            employee=employee, initiated_by=request.user, reason=reason,
            last_working_day=last_working_day,
            notice_period_days=request.data.get('notice_period_days', 0) or 0,
            notes=request.data.get('notes', ''),
        )
        return Response(OffboardingCaseSerializer(case).data, status=201)

    @action(detail=True, methods=['post'], url_path='clear-item/(?P<item_id>[^/.]+)')
    def clear_item(self, request, pk=None, item_id=None):
        """POST /payroll/offboarding-cases/{id}/clear-item/{item_id}/"""
        from django.utils import timezone

        case = self.get_object()
        try:
            item = case.checklist_items.get(id=item_id)
        except ClearanceChecklistItem.DoesNotExist:
            return Response({'error': 'Checklist item not found on this case.'}, status=404)
        item.is_cleared = True
        item.cleared_by = request.user
        item.cleared_at = timezone.now()
        item.save(update_fields=['is_cleared', 'cleared_by', 'cleared_at'])
        if case.status == OffboardingCase.INITIATED:
            case.status = OffboardingCase.IN_PROGRESS
            case.save(update_fields=['status'])
        return Response(ClearanceChecklistItemSerializer(item).data)

    @action(detail=True, methods=['post'])
    def run_final_settlement(self, request, pk=None):
        """POST /payroll/offboarding-cases/{id}/run_final_settlement/"""
        case = self.get_object()
        if case.final_settlement_run_id:
            return Response({'error': 'A final settlement run already exists for this case.'}, status=400)
        run = OffboardingService.run_final_settlement(case, processed_by=request.user)
        return Response(PayrollRunSerializer(run).data, status=201)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """
        POST /payroll/offboarding-cases/{id}/complete/

        Finalizes the case: deactivates the Membership for THIS org only.
        Never touches the User account or other orgs' memberships.
        """
        case = self.get_object()
        if case.status == OffboardingCase.COMPLETED:
            return Response({'error': 'This case is already completed.'}, status=400)
        case = OffboardingService.complete(case, user=request.user)
        return Response(OffboardingCaseSerializer(case).data)

    @action(detail=True, methods=['put', 'patch'])
    def exit_interview(self, request, pk=None):
        """PUT/PATCH /payroll/offboarding-cases/{id}/exit_interview/ — upsert."""
        case = self.get_object()
        interview, _ = ExitInterview.objects.get_or_create(
            organisation=case.organisation, case=case,
        )
        serializer = ExitInterviewSerializer(interview, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(conducted_by=request.user)
        return Response(serializer.data)


class OffboardingChecklistTemplateViewSet(TenantFilterMixin, viewsets.ModelViewSet):
    serializer_class = OffboardingChecklistTemplateSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin, _PlanPayroll]

    def get_queryset(self):
        org = self._get_organisation()
        if not OffboardingChecklistTemplate.objects.filter(organisation=org).exists():
            OffboardingService.seed_checklist_template(org)
        return OffboardingChecklistTemplate.objects.filter(organisation=org)

    def perform_create(self, serializer):
        serializer.save(organisation=self._get_organisation())


# ══════════════════════════════════════════════════════════════════════════════
# HR Analytics (A.6)
# ══════════════════════════════════════════════════════════════════════════════

class HRAnalyticsViewSet(TenantFilterMixin, viewsets.ViewSet):
    """
    Read-only aggregation endpoints. Every query here is a server-side
    .values().annotate() aggregation — never a per-employee Python loop (this
    report class has a documented prior 6.7s N+1 incident on exactly this
    shape elsewhere in the codebase; the same mistake is not repeated here).
    """
    permission_classes = [IsAuthenticated, IsManager, _PlanPayroll]

    # NDPR-motivated: demographic/tenure buckets under this size are suppressed.
    MIN_BUCKET_SIZE = 5

    @staticmethod
    def _suppress_small(counts: dict):
        """Return counts with any bucket < MIN_BUCKET_SIZE replaced by None."""
        return {
            k: (v if v >= HRAnalyticsViewSet.MIN_BUCKET_SIZE else None)
            for k, v in counts.items()
        }

    @action(detail=False, methods=['get'])
    def headcount_turnover(self, request):
        """GET /payroll/hr-analytics/headcount_turnover/?year=<int>"""
        from datetime import date as _date

        from django.db.models.functions import TruncMonth
        from django.db.models import Count

        org = self._get_organisation()
        year = int(request.query_params.get('year') or _date.today().year)

        joiners = (
            Employee.objects.filter(organisation=org, hire_date__year=year)
            .annotate(month=TruncMonth('hire_date'))
            .values('month').annotate(count=Count('id')).order_by('month')
        )
        leavers = (
            Employee.objects.filter(organisation=org, termination_date__year=year)
            .annotate(month=TruncMonth('termination_date'))
            .values('month').annotate(count=Count('id')).order_by('month')
        )
        headcount_start = Employee.all_objects.filter(
            organisation=org, hire_date__lt=_date(year, 1, 1),
        ).exclude(termination_date__lt=_date(year, 1, 1)).count()
        total_joiners = Employee.objects.filter(organisation=org, hire_date__year=year).count()
        total_leavers = Employee.objects.filter(organisation=org, termination_date__year=year).count()
        avg_headcount = headcount_start + (total_joiners - total_leavers) / 2
        attrition_pct = (
            round((total_leavers / avg_headcount) * 100, 2) if avg_headcount > 0 else 0
        )
        def _month_str(m):
            return str(m.date()) if hasattr(m, 'date') else str(m)

        return Response({
            'year': year,
            'headcount_start_of_year': headcount_start,
            'joiners_by_month': [{'month': _month_str(r['month']), 'count': r['count']} for r in joiners],
            'leavers_by_month': [{'month': _month_str(r['month']), 'count': r['count']} for r in leavers],
            'total_joiners': total_joiners,
            'total_leavers': total_leavers,
            'attrition_percent': attrition_pct,
        })

    @action(detail=False, methods=['get'])
    def cost_by_department(self, request):
        """GET /payroll/hr-analytics/cost_by_department/?year=&month="""
        from django.db.models import Count, Sum

        org = self._get_organisation()
        qs = PayslipLine.objects.filter(organisation=org)
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        if year:
            qs = qs.filter(payroll_run__period_year=year)
        if month:
            qs = qs.filter(payroll_run__period_month=month)
        rows = (
            qs.values('employee__department')
            .annotate(total_gross=Sum('gross_salary'), total_net=Sum('net_salary'), headcount=Count('employee', distinct=True))
            .order_by('-total_gross')
        )
        return Response([
            {
                'department': r['employee__department'] or 'Unassigned',
                'total_gross': r['total_gross'], 'total_net': r['total_net'],
                'headcount': r['headcount'],
            }
            for r in rows
        ])

    @action(detail=False, methods=['get'])
    def absence_summary(self, request):
        """GET /payroll/hr-analytics/absence_summary/?year=&month="""
        from django.db.models import Count

        org = self._get_organisation()
        qs = Attendance.objects.filter(organisation=org)
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        if year:
            qs = qs.filter(date__year=year)
        if month:
            qs = qs.filter(date__month=month)
        rows = qs.values('status').annotate(count=Count('id')).order_by('status')
        return Response({r['status']: r['count'] for r in rows})

    @action(detail=False, methods=['get'])
    def tenure_demographics(self, request):
        """
        GET /payroll/hr-analytics/tenure_demographics/

        Tenure and gender/marital-status buckets. Any bucket with fewer than
        MIN_BUCKET_SIZE people is suppressed (returned as null) — NDPR
        requirement, not optional.
        """
        from datetime import date as _date

        from django.db.models import Count

        org = self._get_organisation()
        today = _date.today()

        gender_rows = (
            Employee.objects.filter(organisation=org, is_active=True)
            .values('gender').annotate(count=Count('id'))
        )
        gender_counts = self._suppress_small({r['gender'] or 'unspecified': r['count'] for r in gender_rows})

        marital_rows = (
            Employee.objects.filter(organisation=org, is_active=True)
            .values('marital_status').annotate(count=Count('id'))
        )
        marital_counts = self._suppress_small(
            {r['marital_status'] or 'unspecified': r['count'] for r in marital_rows}
        )

        # Tenure buckets computed via a single annotated query (years-of-service
        # expressed in days, bucketed in Python from the aggregated counts —
        # still one query, no per-employee loop).
        employees = Employee.objects.filter(organisation=org, is_active=True, hire_date__isnull=False).values(
            'hire_date'
        )
        buckets = {'<1yr': 0, '1-3yr': 0, '3-5yr': 0, '5yr+': 0}
        for row in employees:
            days = (today - row['hire_date']).days
            if days < 365:
                buckets['<1yr'] += 1
            elif days < 3 * 365:
                buckets['1-3yr'] += 1
            elif days < 5 * 365:
                buckets['3-5yr'] += 1
            else:
                buckets['5yr+'] += 1
        tenure_buckets = self._suppress_small(buckets)

        return Response({
            'gender': gender_counts,
            'marital_status': marital_counts,
            'tenure_buckets': tenure_buckets,
        })
