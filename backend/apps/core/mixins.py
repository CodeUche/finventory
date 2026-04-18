"""
ViewSet mixins that enforce tenant isolation automatically.

Usage:
    class ProductViewSet(TenantFilterMixin, ModelViewSet):
        ...

All queries are automatically scoped to request.organisation.
This prevents IDOR by making cross-tenant access structurally impossible.

Note on JWT + Middleware order:
    TenantFilterMixin.get_queryset() is called AFTER DRF authentication,
    so request.user is already the JWT-authenticated user here.
    We call resolve_organisation() (phase 2 of tenant resolution) here.
"""

import logging

from apps.core.exceptions import TenantViolationError

logger = logging.getLogger(__name__)


def _safe_str(v):
    """Convert any value to a JSON-safe string for audit diffs."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return str(v)
    try:
        import uuid
        if isinstance(v, uuid.UUID):
            return str(v)
    except Exception:
        pass
    try:
        return str(v)
    except Exception:
        return '(unrepresentable)'


def _changes_from_data(validated_data):
    """Build a simple changes dict from serializer.validated_data (for CREATE events)."""
    changes = {}
    for k, v in validated_data.items():
        changes[k] = {'old': None, 'new': _safe_str(v)}
    return changes


class TenantFilterMixin:
    """
    Mixin that scopes all queries to the current organisation.

    Override `tenant_field` if the model uses a different FK name.
    """

    tenant_field: str = "organisation"

    def _get_organisation(self):
        """
        Resolve (and cache) the organisation for this request.

        Phase 2 of tenant resolution — runs after DRF authentication.

        Also syncs the PostgreSQL session variable ``app.current_org_id`` to
        the *validated* org so that RLS always uses the membership-checked value,
        not just whatever raw UUID the client supplied in the header.

        This fixes two edge-cases:
          1. Fallback resolution: no X-Organisation-ID header → RLSMiddleware
             set SENTINEL, but resolve_organisation() picked the user's first org.
             Without this sync, RLS would block all queries for that request.
          2. Header/validation mismatch: if by any code-path the header and the
             resolved org differ, the DB session is corrected here before any
             queryset is evaluated.
        """
        if getattr(self.request, "organisation", None) is not None:
            # Sync RLS even on cache hit — if a permission class called
            # resolve_organisation() before us (setting request.organisation),
            # the DB session variable may still be SENTINEL (no header path).
            try:
                from apps.core.middleware import _set_org
                _set_org(str(self.request.organisation.id))
            except Exception:
                pass
            return self.request.organisation

        from apps.tenancy.middleware import resolve_organisation
        org = resolve_organisation(self.request)

        if org is None:
            logger.error(
                "TenantFilterMixin: could not resolve organisation for user %s",
                getattr(self.request.user, "id", "anonymous"),
            )
            raise TenantViolationError(
                "Organisation context is missing or you do not have access. "
                "Pass the X-Organisation-ID header."
            )

        # resolve_organisation() already called _sync_rls(), but call _set_org
        # again here as belt-and-suspenders in case the import path differed.
        try:
            from apps.core.middleware import _set_org
            _set_org(str(org.id))
        except Exception:
            pass  # Never let an RLS bookkeeping failure block a request

        return org

    def get_queryset(self):
        org = self._get_organisation()
        qs = super().get_queryset()
        return qs.filter(**{self.tenant_field: org})

    def perform_create(self, serializer):
        """Inject organisation on create + write audit log."""
        org = self._get_organisation()
        serializer.save(**{self.tenant_field: org})
        self._write_audit('create', serializer.instance, _changes_from_data(serializer.validated_data))

    def perform_update(self, serializer):
        """Save + write audit log with field-level diff."""
        instance = serializer.instance
        before = {f: getattr(instance, f, None) for f in serializer.validated_data}
        serializer.save()
        after = {f: getattr(serializer.instance, f, None) for f in serializer.validated_data}
        changes = {
            f: {'old': _safe_str(before[f]), 'new': _safe_str(after[f])}
            for f in before if _safe_str(before[f]) != _safe_str(after[f])
        }
        self._write_audit('update', serializer.instance, changes)

    def perform_destroy(self, instance):
        self._write_audit('delete', instance, {})
        instance.delete()

    def _write_audit(self, action, instance, changes):
        try:
            from apps.core.models import AuditLog
            org = getattr(self.request, 'organisation', None) or self._get_organisation()
            AuditLog.log(
                action=action,
                user=self.request.user if self.request else None,
                organisation=org,
                model_name=instance.__class__.__name__,
                object_id=str(instance.pk),
                object_repr=str(instance),
                changes=changes,
                request=self.request,
            )
        except Exception:
            pass  # Audit failures must never break the main operation


class ExportMixin:
    """
    Adds CSV/XLSX export to any ModelViewSet.

    Append ?dl=csv or ?dl=xlsx to any list endpoint.
    The ViewSet must set `export_fields` (list of (header, field_or_callable) tuples)
    and optionally `export_filename` (base filename without extension).

    Using `dl` instead of `format` avoids DRF's built-in format-suffix content
    negotiation which intercepts `?format=` before the view runs and 404s on
    unknown renderer names like 'csv'.

    Example:
        export_filename = 'invoices'
        export_fields = [
            ('Invoice #', 'invoice_number'),
            ('Customer', lambda obj: obj.customer.name if obj.customer else 'Walk-in'),
            ('Total', 'total'),
        ]
    """

    export_fields: list = []
    export_filename: str = 'export'

    def list(self, request, *args, **kwargs):
        fmt = request.query_params.get('dl', '').lower()
        if fmt in ('csv', 'xlsx'):
            return self._export(request, fmt)
        return super().list(request, *args, **kwargs)

    def _export(self, request, fmt):
        import csv
        import io
        from django.http import HttpResponse

        qs = self.filter_queryset(self.get_queryset())

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.csv"'
            writer = csv.writer(response)
            headers = [h for h, _ in self.export_fields]
            writer.writerow(headers)
            for obj in qs:
                row = []
                for _, field in self.export_fields:
                    if callable(field):
                        row.append(field(obj))
                    else:
                        val = obj
                        for part in field.split('__'):
                            val = getattr(val, part, '') if val is not None else ''
                        row.append(val if val is not None else '')
                writer.writerow(row)
            return response

        # XLSX
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename.capitalize()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1E293B')
        headers = [h for h, _ in self.export_fields]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for obj in qs:
            row = []
            for _, field in self.export_fields:
                if callable(field):
                    row.append(field(obj))
                else:
                    val = obj
                    for part in field.split('__'):
                        val = getattr(val, part, '') if val is not None else ''
                    row.append(str(val) if val is not None else '')
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.xlsx"'
        return response
