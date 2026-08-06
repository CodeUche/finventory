"""
Reporting API views.

All report endpoints are GET-only and accept:
  ?period=today|week|month|year|all|custom   (new — shortcut)
  ?date_from=YYYY-MM-DD                       (used when period=custom)
  ?date_to=YYYY-MM-DD                         (used when period=custom)
  ?format=json|excel|pdf                      (new — triggers file download)

The ?format param is handled by each view via _export_or_json().
"""

import io

from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.permissions import IsManagerOrSuperuser, IsStaff

from .exporters import _org_display_name, _write_report_sheet, dispatch_export, flatten_for_export
from .period_utils import period_label, resolve_period
from .services import ReportService

# ─── Base ─────────────────────────────────────────────────────────────────────


class BaseDateRangeView(APIView):
    """
    Base view that resolves the period / date range from query params and
    provides a helper for optionally returning an export file instead of JSON.
    """

    permission_classes = [IsAuthenticated, IsStaff]

    def get_organisation(self):
        """
        Resolve and return the current organisation.

        IsStaff short-circuits for superusers without calling resolve_organisation(),
        leaving request.organisation=None. This method always resolves it so that
        report queries have the correct org even for superusers.
        """
        if getattr(self.request, "organisation", None) is not None:
            return self.request.organisation
        from apps.tenancy.middleware import resolve_organisation
        return resolve_organisation(self.request)

    def get_date_range(self, request):
        """Return (date_from, date_to) — both None when period='all'."""
        period = request.query_params.get("period", "custom")
        date_from_str = request.query_params.get("date_from")
        date_to_str = request.query_params.get("date_to")
        return resolve_period(period, date_from_str, date_to_str)

    def get_period_label(self, request, date_from=None, date_to=None) -> str:
        period = request.query_params.get("period", "custom")
        return period_label(period, date_from, date_to)

    def get_limit(self, request, default=10, max_limit=100) -> int:
        """Parse and clamp the ?limit= query param to [1, max_limit]."""
        try:
            limit = int(request.query_params.get("limit", default))
        except (TypeError, ValueError):
            limit = default
        return max(1, min(limit, max_limit))

    def _export_or_json(self, request, data, *, headers, row_fn, title, filename_base):
        """
        If ?format=excel|pdf, build an export file and return HttpResponse.
        Otherwise return a DRF Response with *data* as JSON.

        Args:
            data:          The report payload (list or dict).
            headers:       List of column header strings for the export.
            row_fn:        Callable(data) → list[list] converting data to rows.
            title:         Report title for PDF / Excel sheet name.
            filename_base: Filename without extension.
        """
        fmt = request.query_params.get("format", "json").lower()
        if fmt in ("excel", "pdf"):
            date_from, date_to = self.get_date_range(request)
            subtitle = self.get_period_label(request, date_from, date_to)
            rows = row_fn(data)
            response = dispatch_export(
                fmt=fmt,
                headers=headers,
                rows=rows,
                title=title,
                subtitle=subtitle,
                filename_base=filename_base,
                org=self.get_organisation(),
            )
            if response is not None:
                return response
        return Response(data)


# ─── Sales Summary ────────────────────────────────────────────────────────────


class SalesSummaryView(BaseDateRangeView):
    """GET /api/v1/reports/sales/ — Sales summary by period."""

    _HEADERS = ["Period", "Revenue", "Tax Collected", "Discount", "Invoice Count"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        group_by = request.query_params.get("group_by", "day")
        data = list(ReportService.sales_summary(self.get_organisation(), date_from, date_to, group_by))

        def _rows(d):
            return [
                [
                    str(r.get("period", "")),
                    r.get("total_revenue"),
                    r.get("total_tax"),
                    r.get("total_discount"),
                    r.get("invoice_count"),
                ]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Sales Summary",
            filename_base="sales_summary",
        )


# ─── Top Products ─────────────────────────────────────────────────────────────


class TopProductsView(BaseDateRangeView):
    """GET /api/v1/reports/top-products/ — Best-selling products."""

    _HEADERS = ["Product", "SKU", "Units Sold", "Revenue", "COGS", "Gross Profit"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        limit = self.get_limit(request)
        data = ReportService.top_products(self.get_organisation(), date_from, date_to, limit)

        def _rows(d):
            return [
                [r["product_name"], r["product_sku"], r["units_sold"],
                 r["revenue"], r["cogs"], r["gross_profit"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Top Products",
            filename_base="top_products",
        )


# ─── Top Customers ────────────────────────────────────────────────────────────


class TopCustomersView(BaseDateRangeView):
    """GET /api/v1/reports/top-customers/ — Top customers by revenue."""

    _HEADERS = ["Customer", "Code", "Invoice Count", "Revenue"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        limit = self.get_limit(request)
        data = ReportService.top_customers(self.get_organisation(), date_from, date_to, limit)

        def _rows(d):
            return [
                [r["customer_name"], r["customer_code"], r["invoice_count"], r["revenue"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Top Customers",
            filename_base="top_customers",
        )


# ─── Profit & Loss ────────────────────────────────────────────────────────────


class ProfitAndLossView(BaseDateRangeView):
    """GET /api/v1/reports/pnl/ — Profit & Loss statement."""

    _HEADERS = ["Line Item", "Amount (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.profit_and_loss(self.get_organisation(), date_from, date_to)

        def _rows(d):
            rev = d.get("revenue", {})
            return [
                ["Gross Sales", rev.get("gross_sales", 0)],
                ["Tax Collected", rev.get("tax_collected", 0)],
                ["Discounts", rev.get("discounts", 0)],
                ["Cost of Goods Sold", d.get("cost_of_goods_sold", 0)],
                ["Gross Profit", d.get("gross_profit", 0)],
                ["Gross Margin %", d.get("gross_margin_pct", 0)],
                ["Operating Expenses", d.get("operating_expenses", 0)],
                ["Miscellaneous Income", d.get("miscellaneous_income", 0)],
                ["Net Profit", d.get("net_profit", 0)],
                ["Net Margin %", d.get("net_margin_pct", 0)],
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Profit & Loss Statement",
            filename_base="profit_and_loss",
        )


# ─── Expense Breakdown ────────────────────────────────────────────────────────


class ExpenseBreakdownView(BaseDateRangeView):
    """GET /api/v1/reports/expenses/ — Expenses by category."""

    _HEADERS = ["Category", "Total (₦)", "Count"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.expense_breakdown(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [[r["category_name"], r["total"], r["count"]] for r in d]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Expense Breakdown",
            filename_base="expense_breakdown",
        )


# ─── Inventory Valuation ──────────────────────────────────────────────────────


class InventoryValuationView(BaseDateRangeView):
    """GET /api/v1/reports/inventory/ — Current inventory valuation."""

    _HEADERS = ["Product", "SKU", "Warehouse", "Qty", "Unit Cost (₦)", "Total Value (₦)"]

    def get(self, request):
        data = ReportService.inventory_valuation(self.get_organisation())

        def _rows(d):
            return [
                [i["product"], i["sku"], i["warehouse"],
                 i["quantity"], i["unit_cost"], i["total_value"]]
                for i in d.get("items", [])
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Inventory Valuation",
            filename_base="inventory_valuation",
        )


# ─── Cash Flow ────────────────────────────────────────────────────────────────


class CashFlowView(BaseDateRangeView):
    """GET /api/v1/reports/cash-flow/ — Cash flow for period."""

    _HEADERS = ["Line Item", "Amount (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.cash_flow(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [
                ["Cash Inflows", d.get("cash_inflows", 0)],
                ["Cash Outflows", d.get("cash_outflows", 0)],
                ["Net Cash Flow", d.get("net_cash_flow", 0)],
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Cash Flow Statement",
            filename_base="cash_flow",
        )


# ─── AR Aging ─────────────────────────────────────────────────────────────────


class ARAgingView(BaseDateRangeView):
    """GET /api/v1/reports/ar-aging/ — Accounts receivable aging buckets."""

    _HEADERS = ["Invoice #", "Customer", "Amount Due (₦)", "Due Date", "Days Overdue"]

    def get(self, request):
        from datetime import date, datetime

        as_of_str = request.query_params.get("as_of")
        try:
            as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date() if as_of_str else date.today()
        except ValueError:
            as_of = date.today()

        data = ReportService.ar_aging(self.get_organisation(), as_of)

        def _rows(d):
            return [
                [i["invoice_number"], i["customer_name"],
                 i["amount_due"], i["due_date"], i["days_overdue"]]
                for i in d.get("invoices", [])
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Accounts Receivable Aging",
            filename_base="ar_aging",
        )


# ─── AP Aging ─────────────────────────────────────────────────────────────────


class APAgingView(BaseDateRangeView):
    """GET /api/v1/reports/ap-aging/ — Accounts payable aging buckets."""

    _HEADERS = ["Bill #", "Supplier", "Amount Due (₦)", "Due Date", "Days Overdue"]

    def get(self, request):
        from datetime import date, datetime

        as_of_str = request.query_params.get("as_of")
        try:
            as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date() if as_of_str else date.today()
        except ValueError:
            as_of = date.today()

        data = ReportService.ap_aging(self.get_organisation(), as_of)

        def _rows(d):
            return [
                [i["bill_number"], i["supplier_name"],
                 i["amount_due"], i["due_date"], i["days_overdue"]]
                for i in d.get("bills", [])
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Accounts Payable Aging",
            filename_base="ap_aging",
        )


# ─── VAT Summary ─────────────────────────────────────────────────────────────


class VATSummaryView(BaseDateRangeView):
    """GET /api/v1/reports/vat-summary/ — VAT output vs input summary."""

    _HEADERS = ["Line Item", "Amount (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.vat_summary(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [
                ["Output VAT (Collected on Sales)", d.get("output_vat", 0)],
                ["Input VAT (Paid on Bills)", d.get("input_vat", 0)],
                ["Net VAT Payable", d.get("net_vat_payable", 0)],
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="VAT Summary",
            filename_base="vat_summary",
        )


# ─── Sales by Customer ────────────────────────────────────────────────────────


class SalesByCustomerView(BaseDateRangeView):
    """
    GET /api/v1/reports/sales-by-customer/           — all customers with totals
    GET /api/v1/reports/sales-by-customer/?customer_id=<uuid>   — invoices for one customer
    GET /api/v1/reports/sales-by-customer/?customer_id=walk-in  — walk-in invoices
    """

    _SUMMARY_HEADERS = ["Customer", "Code", "Invoices", "Revenue (₦)", "Paid (₦)", "Outstanding (₦)"]
    _DETAIL_HEADERS  = ["Invoice #", "Date", "Status", "Total (₦)", "Paid (₦)", "Due (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        customer_id = request.query_params.get("customer_id")

        if customer_id:
            # Detail mode: invoices for one customer
            cid = None if customer_id == "walk-in" else customer_id
            data = ReportService.customer_invoices(self.get_organisation(), cid, date_from, date_to)

            def _rows(d):
                return [
                    [r["invoice_number"], str(r["issue_date"]), r["status"],
                     r["total_amount"], r["amount_paid"], r["amount_due"]]
                    for r in d
                ]

            return self._export_or_json(
                request, data,
                headers=self._DETAIL_HEADERS,
                row_fn=_rows,
                title="Customer Invoices",
                filename_base="customer_invoices",
            )

        # Summary mode
        data = ReportService.sales_by_customer(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [
                [r["customer_name"], r["customer_code"] or "",
                 r["invoice_count"], r["revenue"],
                 r["amount_paid"], r["amount_outstanding"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._SUMMARY_HEADERS,
            row_fn=_rows,
            title="Sales by Customer",
            filename_base="sales_by_customer",
        )


# ─── Sales by Product ─────────────────────────────────────────────────────────


class SalesByProductView(BaseDateRangeView):
    """
    GET /api/v1/reports/sales-by-product/            — all products with totals
    GET /api/v1/reports/sales-by-product/?product_id=<uuid>   — sale lines for one product
    """

    _SUMMARY_HEADERS = ["Product", "SKU", "Units Sold", "Revenue (₦)", "COGS (₦)", "Gross Profit (₦)"]
    _DETAIL_HEADERS  = ["Invoice #", "Date", "Customer", "Qty", "Unit Price (₦)", "Line Total (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        product_id = request.query_params.get("product_id")

        if product_id:
            data = ReportService.product_sale_lines(self.get_organisation(), product_id, date_from, date_to)

            def _rows(d):
                return [
                    [r["invoice_number"], str(r["issue_date"]), r["customer_name"],
                     r["quantity"], r["unit_price"], r["line_total"]]
                    for r in d
                ]

            return self._export_or_json(
                request, data,
                headers=self._DETAIL_HEADERS,
                row_fn=_rows,
                title="Product Sale Lines",
                filename_base="product_sale_lines",
            )

        data = ReportService.sales_by_product(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [
                [r["product_name"], r["product_sku"] or "",
                 r["units_sold"], r["revenue"], r["cogs"], r["gross_profit"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._SUMMARY_HEADERS,
            row_fn=_rows,
            title="Sales by Product",
            filename_base="sales_by_product",
        )


# ─── Payment Method Breakdown ─────────────────────────────────────────────────


class PaymentMethodsView(BaseDateRangeView):
    """GET /api/v1/reports/payment-methods/ — Revenue by payment method."""

    _HEADERS = ["Method", "Total (₦)", "Transactions"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.payment_methods(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [[r["label"], r["total"], r["count"]] for r in d]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Payment Method Breakdown",
            filename_base="payment_methods",
        )


# ─── Customer Balance ──────────────────────────────────────────────────────


class CustomerBalanceView(BaseDateRangeView):
    """GET /api/v1/reports/customer-balance/ — Outstanding balance snapshot per customer."""

    _HEADERS = ["Customer", "Code", "Outstanding (₦)", "Credit Limit (₦)",
                "Available Credit (₦)", "Last Invoice", "Last Payment"]

    def get(self, request):
        data = ReportService.customer_balance(self.get_organisation())

        def _rows(d):
            return [
                [r["customer_name"], r["customer_code"] or "",
                 r["outstanding_balance"], r["credit_limit"], r["available_credit"],
                 str(r["last_invoice_date"]) if r["last_invoice_date"] else "",
                 str(r["last_payment_date"]) if r["last_payment_date"] else ""]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Customer Balance",
            filename_base="customer_balance",
        )


# ─── Payments by Customer ─────────────────────────────────────────────────


class PaymentsByCustomerView(BaseDateRangeView):
    """GET /api/v1/reports/payments-by-customer/ — SalePayments grouped by customer."""

    _HEADERS = ["Customer", "Total Received (₦)", "Payment Count"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        data = ReportService.payments_by_customer(self.get_organisation(), date_from, date_to)

        def _rows(d):
            return [
                [r["customer_name"], r["total_received"], r["payment_count"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Payments by Customer",
            filename_base="payments_by_customer",
        )


# ─── Customer Payments (drill-down) ───────────────────────────────────────


class CustomerPaymentsView(BaseDateRangeView):
    """GET /api/v1/reports/customer-payments/?customer_id=<uuid> — individual payments for one customer."""

    _HEADERS = ["Date", "Amount (₦)", "Method", "Invoice #", "Reference"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        customer_id = request.query_params.get("customer_id")
        if not customer_id:
            return Response({"error": "customer_id is required"}, status=400)

        data = ReportService.customer_payments(
            self.get_organisation(), customer_id, date_from, date_to
        )

        def _rows(d):
            return [
                [str(r["date"]), r["amount"], r["method"], r["invoice_number"], r["reference"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Customer Payments",
            filename_base="customer_payments",
        )


# ─── Account Statement (GL) ────────────────────────────────────────────────


class AccountStatementView(BaseDateRangeView):
    """GET /api/v1/reports/account-statement/?account_id=<uuid> — GL statement for one account."""

    _HEADERS = ["Date", "Reference", "Description", "Debit (₦)", "Credit (₦)", "Running Balance (₦)"]

    def get(self, request):
        from apps.accounting.models import Account

        date_from, date_to = self.get_date_range(request)
        account_id = request.query_params.get("account_id")
        if not account_id:
            return Response({"error": "account_id is required"}, status=400)

        try:
            data = ReportService.account_statement(
                self.get_organisation(), account_id, date_from, date_to
            )
        except Account.DoesNotExist:
            return Response({"error": "Account not found"}, status=404)

        def _rows(d):
            return [
                [str(r["date"]), r["journal_entry_reference"], r["description"],
                 r["debit"], r["credit"], r["running_balance"]]
                for r in d.get("lines", [])
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Account Statement",
            filename_base="account_statement",
        )


# ─── Customer Invoices (drill-down) ───────────────────────────────────────


class CustomerInvoicesView(BaseDateRangeView):
    """GET /api/v1/reports/customer-invoices/?customer_id=<uuid> — invoices for one customer."""

    _HEADERS = ["Invoice #", "Date", "Status", "Total (₦)", "Paid (₦)", "Due (₦)"]

    def get(self, request):
        date_from, date_to = self.get_date_range(request)
        customer_id = request.query_params.get("customer_id")
        if not customer_id:
            return Response({"error": "customer_id is required"}, status=400)

        cid = None if customer_id == "walk-in" else customer_id
        data = ReportService.customer_invoices(self.get_organisation(), cid, date_from, date_to)

        def _rows(d):
            return [
                [r["invoice_number"], str(r["issue_date"]), r["status"],
                 r["total_amount"], r["amount_paid"], r["amount_due"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Customer Invoices",
            filename_base="customer_invoices",
        )


# ─── Customer Details (master directory) ──────────────────────────────────


class CustomerDetailsView(BaseDateRangeView):
    """GET /api/v1/reports/customer-details/ — Master customer directory."""

    _HEADERS = ["Code", "Name", "Type", "Email", "Phone",
                "Outstanding (₦)", "Credit Limit (₦)", "Total Sales (₦)", "Total Payments (₦)"]

    def get(self, request):
        data = ReportService.customer_details(self.get_organisation())

        def _rows(d):
            return [
                [r["code"], r["name"], r["customer_type"], r["email"], r["phone"],
                 r["outstanding_balance"], r["credit_limit"], r["total_sales"], r["total_payments"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Customer Details",
            filename_base="customer_details",
        )


# ─── Product Details (master directory) ───────────────────────────────────


class ProductDetailsView(BaseDateRangeView):
    """GET /api/v1/reports/product-details/ — Master product directory."""

    _HEADERS = ["SKU", "Name", "Category", "Cost Price (₦)", "Selling Price (₦)",
                "Stock Qty", "Reorder Level", "Margin %"]

    def get(self, request):
        data = ReportService.product_details(self.get_organisation())

        def _rows(d):
            return [
                [r["sku"], r["name"], r["category_name"], r["cost_price"], r["selling_price"],
                 r["stock_quantity"], r["reorder_level"], r["margin_pct"]]
                for r in d
            ]

        return self._export_or_json(
            request, data,
            headers=self._HEADERS,
            row_fn=_rows,
            title="Product Details",
            filename_base="product_details",
        )


# ─── Unified report engine (registry-backed dispatch) ────────────────────────

from . import registry as report_registry  # noqa: E402


class ReportCatalogView(BaseDateRangeView):
    """GET /reports/catalog/ — list registry-backed reports for the reports menu."""

    def get(self, request):
        return Response({"reports": report_registry.catalog()})


class ReportDispatchView(BaseDateRangeView):
    """GET /reports/r/<key>/ — run any registry-backed report by key.

    Accepts the standard period params plus any resolver-specific params
    (e.g. ?account_id= for gl-detail). Returns JSON.
    """

    def get(self, request, key):
        rd = report_registry.get(key)
        if rd is None:
            return Response({"error": f"Unknown report: {key}"}, status=404)
        org = self.get_organisation()
        if org is None:
            return Response({"error": "Organisation not found"}, status=400)
        date_from, date_to = self.get_date_range(request)
        reserved = {"period", "date_from", "date_to", "format"}
        extra = {k: v for k, v in request.query_params.items() if k not in reserved}
        try:
            data = rd.resolver(org, date_from, date_to, **extra)
        except Exception as e:
            return Response({"error": str(e)}, status=422)

        # Generic export: flatten_for_export() knows how to turn every shape
        # the registry's resolvers return (flat rows, bare lists, nested
        # per-account/per-entry sections, summary-figure dicts, ...) into a
        # plain headers/rows/totals table, so export works uniformly across
        # all 30+ reports instead of only the ones that happen to return a
        # top-level `rows` list. See apps/reports/exporters.py for the shape
        # catalogue.
        fmt = request.query_params.get("format", "json").lower()
        if fmt in ("excel", "pdf"):
            flattened = flatten_for_export(data)
            if flattened is not None:
                headers, rows, totals = flattened
                response = dispatch_export(
                    fmt=fmt,
                    headers=headers,
                    rows=rows,
                    title=rd.label,
                    subtitle=self.get_period_label(request, date_from, date_to),
                    filename_base=rd.key,
                    org=org,
                    totals=totals,
                )
                if response is not None:
                    return response

        return Response({
            "key": rd.key, "label": rd.label, "category": rd.category,
            "period_label": self.get_period_label(request, date_from, date_to),
            "data": data,
        })


class ReportBulkExportView(BaseDateRangeView):
    """
    POST /reports/export-bulk/ — export several registry reports at once.

    Body:
        {
          "keys": ["profit-loss", "balance-sheet", ...],  # report keys from /reports/catalog/
          "period": "year", "date_from": "...", "date_to": "...",  # same period params as /reports/r/<key>/
          "combine": true | false,   # true = one .xlsx with one sheet per report; false = a .zip of separate .xlsx files
          "email_to": "someone@example.com"   # optional — if present, email the file instead of returning it
        }

    Runs every requested report's resolver, normalises each one via
    flatten_for_export() (same normaliser the single-report export uses — see
    ReportDispatchView above), and either bundles them into one multi-sheet
    workbook or zips one file per report. A report that fails to run or can't
    be flattened is skipped and reported back in `skipped`, rather than
    failing the whole export — a typo'd key or a resolver that errors on this
    org's data shouldn't block exporting the other 20 reports the user asked for.

    Requires IsManagerOrSuperuser (write-adjacent — this is the same
    permission class the CSV importers use), rather than the read-only
    IsStaff that plain report viewing uses.
    """

    permission_classes = [IsAuthenticated, IsManagerOrSuperuser]

    def post(self, request):
        keys = request.data.get("keys")
        if not isinstance(keys, list) or not keys:
            return Response({"error": "keys must be a non-empty list of report keys"}, status=400)

        org = self.get_organisation()
        if org is None:
            return Response({"error": "Organisation not found"}, status=400)

        # Period comes from the request body here (not query params) since
        # this is a POST with a JSON payload — resolve it the same way
        # BaseDateRangeView.get_date_range()/get_period_label() do for GET.
        from .period_utils import period_label as _period_label, resolve_period as _resolve_period
        period = request.data.get("period", "custom")
        date_from, date_to = _resolve_period(
            period, request.data.get("date_from"), request.data.get("date_to"))
        subtitle = _period_label(period, date_from, date_to)

        combine = bool(request.data.get("combine", True))
        email_to = (request.data.get("email_to") or "").strip()

        # ── Run every requested report and flatten it for export ──────────────
        sheets: list[dict] = []   # {key, label, headers, rows, totals}
        skipped: list[dict] = []  # {key, reason}
        for key in keys:
            rd = report_registry.get(key)
            if rd is None:
                skipped.append({"key": key, "reason": "Unknown report"})
                continue
            try:
                data = rd.resolver(org, date_from, date_to)
            except Exception as e:
                skipped.append({"key": key, "reason": str(e)})
                continue
            flattened = flatten_for_export(data)
            if flattened is None:
                skipped.append({"key": key, "reason": "This report has no tabular export"})
                continue
            headers, rows, totals = flattened
            sheets.append({"key": rd.key, "label": rd.label, "headers": headers, "rows": rows, "totals": totals})

        if not sheets:
            return Response({"error": "None of the requested reports could be exported", "skipped": skipped}, status=422)

        # ── Build the file(s) ───────────────────────────────────────────────────
        org_name = _org_display_name(org)
        if combine:
            file_bytes = self._build_workbook(sheets, org_name, subtitle)
            filename = "audity-reports.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            file_bytes = self._build_zip(sheets, org_name, subtitle)
            filename = "audity-reports.zip"
            content_type = "application/zip"

        # ── Email or download ────────────────────────────────────────────────────
        if email_to:
            return self._email_file(request, org, email_to, filename, content_type, file_bytes, skipped)

        response = HttpResponse(file_bytes, content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        if skipped:
            # Surface partial-failure info without failing the download —
            # the browser ignores unknown response headers, the frontend can
            # read this one to show a "N reports skipped" toast after saving.
            import json as _json
            response["X-Reports-Skipped"] = _json.dumps(skipped)[:4000]
        return response

    @staticmethod
    def _build_workbook(sheets: "list[dict]", org_name: str, subtitle: str) -> bytes:
        """One workbook, one sheet per report — see _write_report_sheet() in exporters.py."""
        import openpyxl

        wb = openpyxl.Workbook()
        used_names: set = set()
        for i, s in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = _unique_sheet_name(s["label"], used_names)
            _write_report_sheet(
                ws, s["headers"], s["rows"],
                title=s["label"], subtitle=subtitle, org_name=org_name, totals=s["totals"],
            )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _build_zip(sheets: "list[dict]", org_name: str, subtitle: str) -> bytes:
        """One .xlsx per report, zipped together (stdlib zipfile — no new dependency).
        Each file is built the same way _build_workbook() builds a sheet (via
        _write_report_sheet, passing org_name directly) rather than through
        export_excel()'s `org=` object param, so both bulk-export modes share
        one code path for the header block/totals row instead of two."""
        import zipfile

        import openpyxl

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            used_names: set = set()
            for s in sheets:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = s["label"][:31]
                _write_report_sheet(
                    ws, s["headers"], s["rows"],
                    title=s["label"], subtitle=subtitle, org_name=org_name, totals=s["totals"],
                )
                sheet_buf = io.BytesIO()
                wb.save(sheet_buf)
                arcname = _unique_sheet_name(s["key"], used_names, ext=".xlsx")
                zf.writestr(arcname, sheet_buf.getvalue())
        return buf.getvalue()

    def _email_file(self, request, org, email_to, filename, content_type, file_bytes, skipped):
        """Send the export as an email attachment via the org's configured SMTP.
        Mirrors QuoteViewSet.send_email (apps/quotes/views.py) — same config
        lookup, same MIME construction, same error responses — so failures
        read consistently across the app regardless of which feature sent them."""
        import smtplib
        import ssl as _ssl
        from email import encoders
        from email.mime.base import MIMEBase
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        try:
            email_cfg = org.email_config
            if not email_cfg.is_active:
                return Response({"error": "Email is not configured. Go to Settings → Email."}, status=422)
        except Exception:
            return Response({"error": "Email is not configured. Go to Settings → Email."}, status=422)

        from_name = email_cfg.from_name or org.name
        from_email = email_cfg.from_email or email_cfg.smtp_username

        msg = MIMEMultipart("mixed")
        msg["Subject"] = f"Reports export from {from_name}"
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = email_to

        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(
            f"<p>Attached: {filename} — {len(file_bytes):,} bytes, "
            f"{'a workbook with one sheet per report' if filename.endswith('.xlsx') else 'a zip of one file per report'}.</p>",
            "html",
        ))
        msg.attach(alt)

        part = MIMEBase(*content_type.split("/", 1))
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)

        try:
            ctx = _ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = _ssl.CERT_NONE
            if email_cfg.use_tls:
                conn = smtplib.SMTP(email_cfg.smtp_host, email_cfg.smtp_port, timeout=30)
                conn.ehlo(); conn.starttls(context=ctx); conn.ehlo()
            else:
                conn = smtplib.SMTP_SSL(email_cfg.smtp_host, email_cfg.smtp_port, timeout=30, context=ctx)
                conn.ehlo()
            conn.login(email_cfg.smtp_username, email_cfg.smtp_password)
            conn.sendmail(from_email, [email_to], msg.as_string())
            try:
                conn.quit()
            except Exception:
                pass
            return Response({"message": f"Reports sent to {email_to}", "skipped": skipped})
        except smtplib.SMTPAuthenticationError:
            return Response({"error": "SMTP authentication failed. Check your username and password in Settings → Email."}, status=422)
        except Exception:
            import logging as _log
            _log.getLogger(__name__).exception("Bulk report email send failed")
            return Response({"error": "Failed to send email. Please check your SMTP settings."}, status=422)


def _unique_sheet_name(label: str, used: set, ext: str = "") -> str:
    """Excel sheet names (and our zip entry names, for tidiness) must be <=31
    chars and unique within the file — truncate and de-dupe with a numeric
    suffix if two report labels collide after truncation."""
    max_len = 31 - len(ext) if ext else 31
    base = (label or "Report")[:max_len]
    name = f"{base}{ext}"
    n = 2
    while name.lower() in used:
        suffix = f" ({n}){ext}"
        name = f"{base[:max_len - len(suffix) + len(ext)]}{suffix}"
        n += 1
    used.add(name.lower())
    return name
